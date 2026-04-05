# =============================================================================
# PLANIFICACIÓN CARTOGRÁFICA — STREAMLIT v5.1
# INEC · Zonal Litoral · Autores: Franklin López, Carlos Quinto
#
# CAMBIOS v5.1 (sobre v5):
#
# ── NUEVO: Edición manual post-planificación ────────────────────────────────
#   Nuevo tab "✏️ Edición Manual" que permite:
#     - Reasignar UPMs/manzanas a otro equipo, jornada o encuestador
#     - Mover rangos de días operativos
#     - Filtrado por equipo/jornada para edición rápida
#     - Resumen de cambios antes de confirmar
#     - Botón para recalcular métricas tras edición
#
# ── NUEVO: Tooltips en sliders ──────────────────────────────────────────────
#   Todos los sliders del sidebar ahora muestran ayuda contextual
#   al pasar el ratón (parámetro help= de Streamlit).
# =============================================================================

import streamlit as st
import pandas as pd
import numpy as np
import geopandas as gpd
import folium
import pyogrio
from streamlit_folium import st_folium
import plotly.express as px
import plotly.graph_objects as go
import tempfile, os, warnings, io
from datetime import date, timedelta
import osmnx as ox
import networkx as nx
from networkx.algorithms import approximation
from pyproj import Transformer
from sklearn.cluster import KMeans
from sklearn.metrics import silhouette_score
from sklearn.neighbors import BallTree
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')

# ── PAGE CONFIG ───────────────────────────────
st.set_page_config(page_title="Planificación",
                   page_icon="🗺️", layout="wide",
                   initial_sidebar_state="expanded")

# ── CSS ───────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600&display=swap');
html,body,[class*="css"]{font-family:'IBM Plex Sans',sans-serif}
[data-testid="stSidebar"]{background:#0c0f1a;border-right:1px solid #1e2540}
[data-testid="stSidebar"] *{color:#d0d8e8 !important}
.hdr{background:linear-gradient(135deg,#071e3d,#0d3b6e 60%,#0a2a52);
     border-radius:12px;padding:24px 32px;margin-bottom:20px;
     border-left:5px solid #2e86de;position:relative;overflow:hidden}
.hdr::after{content:"INEC";position:absolute;right:24px;top:50%;
            transform:translateY(-50%);font-family:'IBM Plex Mono',monospace;
            font-size:76px;font-weight:600;color:rgba(255,255,255,.04);letter-spacing:6px}
.hdr h1{color:#fff!important;font-size:18px!important;font-weight:600!important;
        margin:0 0 3px!important;font-family:'IBM Plex Mono',monospace!important}
.hdr p{color:#7eb3d8!important;font-size:12px!important;margin:0!important}
.kcard{background:#111827;border:1px solid #1f2d45;border-radius:10px;
       padding:14px 16px;text-align:center;transition:border-color .2s}
.kcard:hover{border-color:#2e86de}
.kcard .v{font-family:'IBM Plex Mono',monospace;font-size:24px;font-weight:600;
          color:#2e86de;line-height:1}
.kcard .l{font-size:10px;color:#7a8fa6;margin-top:4px;text-transform:uppercase;letter-spacing:.5px}
.kcard .s{font-size:10px;color:#4a6070;margin-top:2px}
.step{display:inline-block;background:#0d2035;color:#2e86de;border:1px solid #1a4060;
      border-radius:4px;padding:2px 7px;font-family:'IBM Plex Mono',monospace;
      font-size:10px;font-weight:600;letter-spacing:1px;margin-bottom:6px}
.stitle{font-family:'IBM Plex Mono',monospace;font-size:11px;font-weight:600;color:#2e86de;
        text-transform:uppercase;letter-spacing:1px;border-bottom:1px solid #1f2d45;
        padding-bottom:7px;margin:18px 0 12px}
.ibox{background:#0a1f35;border:1px solid #143050;border-left:3px solid #2e86de;
      border-radius:7px;padding:11px 15px;margin:9px 0;font-size:13px;color:#7eb3d8}
.wbox{background:#1a1400;border:1px solid #3a2800;border-left:3px solid #f39c12;
      border-radius:7px;padding:11px 15px;margin:9px 0;font-size:13px;color:#c9a227}
.bcard{background:#1a0d2e;border:1px solid #3d1a6e;border-left:3px solid #9b59b6;
       border-radius:7px;padding:13px 16px;margin:9px 0}
.pill-ok{display:inline-block;background:#0a2e1a;color:#27ae60;border:1px solid #1a5e35;
         border-radius:20px;padding:2px 9px;font-size:11px;
         font-family:'IBM Plex Mono',monospace;font-weight:600}
.pill-w{display:inline-block;background:#1a1500;color:#e67e22;border:1px solid #5a3c00;
        border-radius:20px;padding:2px 9px;font-size:11px;
        font-family:'IBM Plex Mono',monospace;font-weight:600}
.eq-card{background:#0d1520;border:1px solid #1f2d45;border-radius:9px;
         padding:14px 16px;text-align:center;transition:border-color .2s}
.eq-card:hover{border-color:#2e86de}
.pi-form{background:#0d1520;border:1px solid #1f2d45;border-radius:8px;
         padding:16px;margin-bottom:12px}
.balance-box{background:#071a10;border:1px solid #0d4020;border-left:3px solid #27ae60;
             border-radius:7px;padding:11px 15px;margin:9px 0;font-size:12px;color:#5dca8a}
.jplan-ok  {background:#061a0e;border:1px solid #0d4020;border-left:3px solid #27ae60;
             border-radius:6px;padding:8px 14px;margin:4px 0;font-size:12px;color:#5dca8a}
.jplan-mv  {background:#1a1200;border:1px solid #3a2800;border-left:3px solid #f39c12;
             border-radius:6px;padding:8px 14px;margin:4px 0;font-size:12px;color:#c9a227}
.jplan-can {background:#1a0800;border:1px solid #3a1000;border-left:3px solid #e74c3c;
             border-radius:6px;padding:8px 14px;margin:4px 0;font-size:12px;color:#e07060}
.edit-card {background:#0d1a2e;border:1px solid #1a3050;border-left:3px solid #e67e22;
            border-radius:7px;padding:13px 16px;margin:9px 0;font-size:13px;color:#c9a227}
.edit-ok   {background:#061a0e;border:1px solid #0d4020;border-left:3px solid #27ae60;
            border-radius:7px;padding:11px 15px;margin:9px 0;font-size:13px;color:#5dca8a}
.edit-count{display:inline-block;background:#1a0d2e;color:#9b59b6;border:1px solid #3d1a6e;
            border-radius:20px;padding:3px 12px;font-size:12px;
            font-family:'IBM Plex Mono',monospace;font-weight:600;margin-right:6px}
</style>
""", unsafe_allow_html=True)

# ── CONSTANTES ────────────────────────────────
BASE_LAT = -2.145825935522539
BASE_LON = -79.89383956329586
PRO_GYE  = "09"
CAN_GYE  = "01"

MESES_CAL = {
    1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
    7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"
}

COLORES  = ['#e74c3c','#2e86de','#27ae60','#f39c12','#9b59b6',
            '#1abc9c','#e67e22','#e91e63']

# ── HELPERS ───────────────────────────────────
def cv_pct(s):
    m = s.mean()
    return float(s.std()/m*100) if m > 0 else 0.0

def utm_to_wgs84(df):
    t = Transformer.from_crs("epsg:32717","epsg:4326",always_xy=True)
    lons,lats = t.transform(df["x"].values,df["y"].values)
    df = df.copy(); df["lon"]=lons; df["lat"]=lats
    return df

def parse_codigo(codigo):
    c = str(codigo).strip()
    r = {'prov':'','canton':'','ciudad_parroq':'','zona':'','sector':'','man':''}
    if len(c)>=6:  r['prov']=c[:2]; r['canton']=c[2:4]; r['ciudad_parroq']=c[4:6]
    if len(c)>=9:  r['zona']=c[6:9]
    if len(c)>=12: r['sector']=c[9:12]
    if len(c)>=15: r['man']=c[12:15]
    return r

def normalizar_codigo(valor, ancho=None):
    if pd.isna(valor):
        return ''
    s = str(valor).strip()
    if s.endswith('.0'):
        s = s[:-2]
    s = ''.join(ch for ch in s if ch.isdigit())
    if not s:
        return ''
    if ancho:
        s = s.zfill(ancho)[-ancho:]
    return s

def detectar_columnas_catalogo(df_cat):
    cols = {str(c).strip().upper(): c for c in df_cat.columns}
    def pick(*candidatas):
        for cand in candidatas:
            if cand in cols:
                return cols[cand]
        return None
    return {
        'parroquia_cod': pick('DPA_PARROQ', 'COD_PARROQUIA', 'PARROQUIA_CODIGO'),
        'parroquia_nom': pick('DPA_DESPAR', 'DESC_PARROQUIA', 'PARROQUIA'),
        'canton_cod'   : pick('DPA_CANTON', 'COD_CANTON', 'CANTON_CODIGO'),
        'canton_nom'   : pick('DPA_DESCAN', 'DESC_CANTON', 'CANTON'),
        'prov_cod'     : pick('DPA_PROVIN', 'COD_PROVINCIA', 'PROVINCIA_CODIGO'),
        'prov_nom'     : pick('DPA_DESPRO', 'DESC_PROVINCIA', 'PROVINCIA'),
        'tipo_txt'     : pick('TXT', 'TIPO', 'TIPO_SECTOR', 'CLASE'),
        'fcode'        : pick('FCODE', 'COD_FCODE', 'CODIGO_FCODE')
    }

def cargar_catalogo_territorial(file_obj):
    nombre = getattr(file_obj, 'name', '').lower()
    if nombre.endswith('.csv'):
        df_cat = pd.read_csv(file_obj)
    else:
        df_cat = pd.read_excel(file_obj)
    df_cat.columns = [str(c).strip() for c in df_cat.columns]
    return df_cat

def preparar_lookup_territorial(df_cat):
    if df_cat is None or len(df_cat) == 0:
        return {}, {}
    cols = detectar_columnas_catalogo(df_cat)
    cod_col = cols.get('parroquia_cod')
    if not cod_col:
        return {}, cols
    work = df_cat.copy()
    work['__parroq_cod__'] = work[cod_col].apply(lambda v: normalizar_codigo(v, 6))
    work = work[work['__parroq_cod__'] != ''].drop_duplicates('__parroq_cod__', keep='first')
    lookup = {}
    for _, row in work.iterrows():
        cod = row['__parroq_cod__']
        lookup[cod] = {
            'provincia_codigo' : normalizar_codigo(row[cols['prov_cod']], 2) if cols.get('prov_cod') else cod[:2],
            'provincia_nombre' : str(row[cols['prov_nom']]).strip() if cols.get('prov_nom') and pd.notna(row[cols['prov_nom']]) else '',
            'canton_codigo'    : normalizar_codigo(row[cols['canton_cod']], 4) if cols.get('canton_cod') else cod[:4],
            'canton_nombre'    : str(row[cols['canton_nom']]).strip() if cols.get('canton_nom') and pd.notna(row[cols['canton_nom']]) else '',
            'parroquia_codigo' : cod,
            'parroquia_nombre' : str(row[cols['parroquia_nom']]).strip() if cols.get('parroquia_nom') and pd.notna(row[cols['parroquia_nom']]) else '',
            'tipo_txt'         : str(row[cols['tipo_txt']]).strip() if cols.get('tipo_txt') and pd.notna(row[cols['tipo_txt']]) else '',
            'fcode'            : str(row[cols['fcode']]).strip() if cols.get('fcode') and pd.notna(row[cols['fcode']]) else ''
        }
    return lookup, cols

def enriquecer_plan_con_catalogo(df_plan, catalogo_lookup):
    if df_plan is None or len(df_plan) == 0:
        return df_plan
    if not catalogo_lookup:
        return df_plan.copy()
    df_out = df_plan.copy()
    provs, cants, parroqs, tipos = [], [], [], []
    for _, row in df_out.iterrows():
        partes = parse_codigo(row.get('id_entidad', ''))
        cod_parr = f"{partes['prov']}{partes['canton']}{partes['ciudad_parroq']}"
        geo = catalogo_lookup.get(cod_parr, {})
        provs.append(geo.get('provincia_nombre', ''))
        cants.append(geo.get('canton_nombre', ''))
        parroqs.append(geo.get('parroquia_nombre', ''))
        tipos.append(geo.get('tipo_txt', ''))
    df_out['provincia_nombre'] = provs
    df_out['canton_nombre'] = cants
    df_out['parroquia_nombre'] = parroqs
    df_out['tipo_asentamiento'] = tipos
    return df_out

def jornada_num_desde_mes(mes_operativo, mes_inicio_cal):
    mes_cal  = ((mes_inicio_cal - 1 + mes_operativo - 1) % 12) + 1
    j1_num   = (mes_operativo - 1) * 2 + 1
    j2_num   = j1_num + 1
    return j1_num, j2_num, MESES_CAL[mes_cal]

def cargar_gpkg(path, dissolve_upm=True):
    capas = pyogrio.list_layers(path)
    if len(capas) == 1:
        gdf = gpd.read_file(path, layer=capas[0][0])
        col_map = {
            '1_mes_cart':'mes','viv_total':'viv','1_zonal':'zonal',
            '1_id_upm':'upm','ManSec':'id_entidad'
        }
        gdf = gdf.rename(columns={k:v for k,v in col_map.items() if k in gdf.columns})
        if 'id_entidad' in gdf.columns:
            gdf['tipo_entidad'] = gdf['id_entidad'].astype(str).apply(
                lambda x: 'sec' if '999' in x else 'man')
        gdf_u = gdf.to_crs(epsg=32717)
        gdf_u['geometry'] = gdf_u.geometry.representative_point()
        gdf_u['x'] = gdf_u.geometry.x
        gdf_u['y'] = gdf_u.geometry.y
        if 'pro' in gdf_u.columns: gdf_u['pro_x'] = gdf_u['pro']
        if 'can' in gdf_u.columns: gdf_u['can_x'] = gdf_u['can']
        if 'mes' in gdf_u.columns: gdf_u['mes'] = pd.to_numeric(gdf_u['mes'],errors='coerce')
        if dissolve_upm and 'upm' in gdf_u.columns:
            agg = {'viv':'sum','mes':'first','x':'first','y':'first','tipo_entidad':'first'}
            if 'pro_x' in gdf_u.columns: agg['pro_x']='first'
            if 'can_x' in gdf_u.columns: agg['can_x']='first'
            gdf_f = gdf_u.groupby('upm').agg(agg).reset_index()
            gdf_f['id_entidad'] = gdf_f['upm']
            gdf_f['tipo_entidad'] = gdf_f['tipo_entidad'].apply(lambda t: f"{t}_upm")
        else:
            gdf_f = gdf_u
        return utm_to_wgs84(gdf_f)
    else:
        man  = gpd.read_file(path,layer=capas[0][0])
        disp = gpd.read_file(path,layer=capas[1][0])
        man  = man[man['zonal']=='LITORAL']
        disp = disp[disp['zonal']=='LITORAL']
        man_u = man.to_crs(epsg=32717); dis_u = disp.to_crs(epsg=32717)
        if dissolve_upm:
            def _d(gdf,tipo):
                d = gdf.dissolve(by='upm',aggfunc={'mes':'first','viv':'sum'})
                d['geometry'] = d.geometry.representative_point()
                o = d[['mes','viv']].copy()
                o['id_entidad']=d.index; o['upm']=d.index; o['tipo_entidad']=tipo
                o['x']=d.geometry.x; o['y']=d.geometry.y
                if 'mes' in o.columns: o['mes']=pd.to_numeric(o['mes'],errors='coerce')
                return o[['id_entidad','upm','mes','viv','x','y','tipo_entidad']]
            ms=_d(man_u,'man_upm'); ds=_d(dis_u,'sec_upm')
        else:
            for g in [man_u,dis_u]:
                g['geometry']=g.geometry.representative_point()
                g['x']=g.geometry.x; g['y']=g.geometry.y
            ms=man_u[['man','upm','mes','viv','x','y']].rename(columns={'man':'id_entidad'})
            ms['tipo_entidad']='man'
            ds=dis_u[['sec','upm','mes','viv','x','y']].rename(columns={'sec':'id_entidad'})
            ds['tipo_entidad']='sec'
            ms['pro_x']=ms['id_entidad'].astype(str).str[:2]
            ms['can_x']=ms['id_entidad'].astype(str).str[2:4]
            for df_t in [ms,ds]:
                if 'mes' in df_t.columns: df_t['mes']=pd.to_numeric(df_t['mes'],errors='coerce')
        data=pd.concat([ms,ds],ignore_index=True)
        if not dissolve_upm: data=data.drop_duplicates(subset=['id_entidad','upm'],keep='first')
        return utm_to_wgs84(data)


# ══════════════════════════════════════════════════════════════════════════════
#  CLUSTERING BALANCEADO
# ══════════════════════════════════════════════════════════════════════════════

def clustering_balanceado(df, n_clusters, cv_objetivo=0.10, max_iter=300, k_vecinos=8):
    coords = df[['x','y']].values.astype(float)
    cargas = df['carga_pond'].values.astype(float)
    n      = len(df)

    km     = KMeans(n_clusters=n_clusters,init='k-means++',n_init=20,
                    max_iter=500,random_state=42)
    labels = km.fit_predict(coords).copy()

    def cluster_sums():
        return np.array([cargas[labels==c].sum() for c in range(n_clusters)])

    def centroides_actuales():
        return np.array([
            coords[labels==c].mean(axis=0) if (labels==c).sum()>0 else np.zeros(2)
            for c in range(n_clusters)
        ])

    cv_ini = cv_pct(pd.Series(cluster_sums()))
    log    = [{'iter':0,'cv':cv_ini,'modo':'inicial'}]

    tree,_  = BallTree(coords,leaf_size=40), None
    _,nbr   = tree.query(coords,k=min(k_vecinos+1,n))

    no_mejora_frontera = 0
    cv_history         = [cv_ini]

    for it in range(1, max_iter+1):
        sums = cluster_sums()
        cv   = cv_pct(pd.Series(sums))

        if cv <= cv_objetivo*100:
            log.append({'iter':it,'cv':cv,'modo':'objetivo alcanzado ✓'})
            break

        cv_history.append(cv)
        if len(cv_history) >= 10:
            mejora_reciente = cv_history[-10] - cv_history[-1]
            if mejora_reciente < 0.1:
                log.append({'iter':it,'cv':cv,'modo':'plateau — equilibrio estable'})
                break

        orden_pesados  = np.argsort(sums)[::-1]
        orden_livianos = np.argsort(sums)
        mejora         = False

        if no_mejora_frontera < 3:
            for ci_p in orden_pesados[:3]:
                for ci_l in orden_livianos[:3]:
                    if ci_p==ci_l: continue
                    mask_p = np.where(labels==ci_p)[0]
                    if len(mask_p)==0: continue
                    mejor_cv,mejor_idx = cv,-1
                    for idx in mask_p:
                        if not any(labels[v]==ci_l for v in nbr[idx] if v!=idx): continue
                        labels[idx]=ci_l
                        cv_n=cv_pct(pd.Series(cluster_sums()))
                        labels[idx]=ci_p
                        if cv_n<mejor_cv: mejor_cv,mejor_idx=cv_n,idx
                    if mejor_idx>=0:
                        labels[mejor_idx]=ci_l
                        log.append({'iter':it,'cv':mejor_cv,'modo':'frontera'})
                        mejora=True; no_mejora_frontera=0; break
                if mejora: break
            if not mejora: no_mejora_frontera+=1

        if not mejora:
            cents  = centroides_actuales()
            ci_p   = orden_pesados[0]
            mask_p = np.where(labels==ci_p)[0]
            mejor_cv_g,mejor_idx_g,mejor_dest = cv,-1,-1
            for ci_l in orden_livianos[:2]:
                if ci_p==ci_l or len(mask_p)==0: continue
                dists = np.linalg.norm(coords[mask_p]-cents[ci_l],axis=1)
                cands = mask_p[np.argsort(dists)[:10]]
                for idx in cands:
                    labels[idx]=ci_l
                    cv_n=cv_pct(pd.Series(cluster_sums()))
                    labels[idx]=ci_p
                    if cv_n<mejor_cv_g: mejor_cv_g,mejor_idx_g,mejor_dest=cv_n,idx,ci_l
            if mejor_idx_g>=0 and mejor_cv_g<cv:
                labels[mejor_idx_g]=mejor_dest
                log.append({'iter':it,'cv':mejor_cv_g,'modo':'global'})
                mejora=True; no_mejora_frontera=0
            else:
                log.append({'iter':it,'cv':cv,'modo':'sin mejora'})
                if sum(1 for l in log[-6:] if l.get('modo')=='sin mejora')>=5:
                    break

    cv_fin = cv_pct(pd.Series(cluster_sums()))
    log.append({'iter':len(log),'cv':cv_fin,'modo':'final'})
    return labels, log, cv_ini, cv_fin


# ══════════════════════════════════════════════════════════════════════════════
#  NEAREST-NEIGHBOR
# ══════════════════════════════════════════════════════════════════════════════

def nearest_neighbor_order(points_xy, start_xy=None):
    n = len(points_xy)
    if n==0: return []
    if n==1: return [0]
    visited=[False]*n
    if start_xy is not None:
        cur=int(np.argmin(np.linalg.norm(points_xy-start_xy,axis=1)))
    else:
        cur=0
    order=[cur]; visited[cur]=True
    for _ in range(n-1):
        best_d,best_j=np.inf,-1
        px,py=points_xy[cur]
        for j in range(n):
            if not visited[j]:
                d=(points_xy[j,0]-px)**2+(points_xy[j,1]-py)**2
                if d<best_d: best_d,best_j=d,j
        cur=best_j; order.append(cur); visited[cur]=True
    return order


# ══════════════════════════════════════════════════════════════════════════════
#  ASIGNACIÓN ENCUESTADORES + DÍAS
# ══════════════════════════════════════════════════════════════════════════════

def asignar_encuestadores_y_dias(df_grp, n_enc, dias_tot, viv_min, viv_max,
                                  inicio_dia=1):
    target = (viv_min + viv_max) / 2.0
    ultimo = inicio_dia + dias_tot - 1

    df_g   = df_grp.copy()
    idx_orig = df_g.index.tolist()
    n_rows = len(df_g)

    cargas = df_g['carga_pond'].values.astype(float)
    viviendas = df_g['viv'].values.astype(float)
    coords = df_g[['x','y']].values.astype(float)

    orden_bp = np.argsort(cargas)[::-1]
    enc_acum = np.zeros(n_enc)
    enc_asig = np.zeros(n_rows, dtype=int)

    for pos in orden_bp:
        em = int(np.argmin(enc_acum))
        enc_asig[pos] = em + 1
        enc_acum[em] += cargas[pos]

    geo_order = []
    for enc_id in range(1, n_enc+1):
        pos_enc = np.where(enc_asig == enc_id)[0]
        if len(pos_enc)==0: continue
        sub_coords = coords[pos_enc]
        centroide  = sub_coords.mean(axis=0)
        nn         = nearest_neighbor_order(sub_coords, start_xy=centroide)
        for nn_pos in nn:
            geo_order.append((pos_enc[nn_pos], enc_id))

    dias_range = list(range(inicio_dia, ultimo+1))
    calendario = {e:{d:0.0 for d in dias_range} for e in range(1,n_enc+1)}
    cursor     = {e:inicio_dia for e in range(1,n_enc+1)}

    dia_ini_arr = np.full(n_rows, inicio_dia, dtype=int)
    dia_fin_arr = np.full(n_rows, inicio_dia, dtype=int)

    for pos, enc_id in geo_order:
        viv_m = max(0.0, viviendas[pos])
        cal   = calendario[enc_id]
        cur   = cursor[enc_id]

        if viv_m > viv_max:
            dias_m = max(1, int(np.ceil(viv_m / target)))
            dias_m = min(dias_m, dias_tot)
            bloque = None
            for d_s in range(cur, ultimo - dias_m + 2):
                if all(cal.get(d, target) < target for d in range(d_s, d_s+dias_m)):
                    bloque = d_s; break
            if bloque is None:
                bloque = max(inicio_dia, min(cur, ultimo - dias_m + 1))
            d_ini = bloque
            d_fin = min(d_ini + dias_m - 1, ultimo)
            vpd   = viv_m / max(1, d_fin - d_ini + 1)
            for dd in range(d_ini, d_fin+1):
                cal[dd] = cal.get(dd, 0.0) + vpd
            cursor[enc_id] = d_fin + 1
        else:
            dia_asig = None
            for d in range(cur, ultimo+1):
                if cal.get(d, 0.0) < target:
                    dia_asig = d; break
            if dia_asig is None: dia_asig = ultimo
            cal[dia_asig] = cal.get(dia_asig, 0.0) + viv_m
            d_ini = dia_asig; d_fin = dia_asig
            if cal[dia_asig] >= target and cursor[enc_id] == dia_asig:
                cursor[enc_id] = min(dia_asig+1, ultimo)

        d_ini = max(inicio_dia, min(d_ini, ultimo))
        d_fin = max(d_ini,      min(d_fin, ultimo))
        dia_ini_arr[pos] = d_ini
        dia_fin_arr[pos] = d_fin

    df_g['encuestador']   = enc_asig
    df_g['dia_inicio']    = dia_ini_arr
    df_g['dia_fin']       = dia_fin_arr
    df_g['dia_operativo'] = dia_ini_arr
    return df_g


# ══════════════════════════════════════════════════════════════════════════════
#  PLANIFICACIÓN DE JORNADAS
# ══════════════════════════════════════════════════════════════════════════════

ESTADO_OK  = "✅ Planificada"
ESTADO_MV  = "🔀 Trasladada"
ESTADO_CAN = "❌ Cancelada"

def construir_calendario_jornadas(total_meses, mes_inicio_cal, config_jornadas):
    jornadas = []
    for mes in range(1, total_meses+1):
        j1_n, j2_n, mes_nombre = jornada_num_desde_mes(mes, mes_inicio_cal)
        for jn in [j1_n, j2_n]:
            cfg = config_jornadas.get(jn, {})
            jornadas.append({
                'jornada_num': jn,
                'mes_op': mes,
                'mes_nombre': mes_nombre,
                'mitad': 1 if jn==j1_n else 2,
                'estado': cfg.get('estado', ESTADO_OK),
                'fecha': cfg.get('fecha', None),
                'trasladada_a': cfg.get('trasladada_a', None),
            })
    return jornadas


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def generar_excel(df_plan, eq_cfg, personal_info,
                  jornadas_activas, dias_op, catalogo_lookup=None):
    catalogo_lookup = catalogo_lookup or {}
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    AZ_OSCURO="0D3B6E"; AZ_MEDIO="1A5276"; AZ_CLARO="D6EAF8"
    VRD_CHECK="D5F5E3"; BLANCO="FFFFFF"
    ENC_PALETAS = [
        {"par":"DBEAFE","impar":"EFF6FF","subtot":"BFDBFE","hdr":"1D4ED8"},
        {"par":"D1FAE5","impar":"ECFDF5","subtot":"A7F3D0","hdr":"065F46"},
        {"par":"FEF9C3","impar":"FEFCE8","subtot":"FDE68A","hdr":"854D0E"},
        {"par":"FCE7F3","impar":"FDF4FF","subtot":"F9A8D4","hdr":"831843"},
        {"par":"FFE4E6","impar":"FFF1F2","subtot":"FECACA","hdr":"9F1239"},
        {"par":"E0E7FF","impar":"EEF2FF","subtot":"C7D2FE","hdr":"3730A3"},
    ]

    def sc(cell,bold=False,bg=None,fg="000000",ha="left",sz=9,brd=False,wrap=False):
        cell.font=Font(bold=bold,size=sz,color=fg)
        cell.alignment=Alignment(horizontal=ha,vertical="center",wrap_text=wrap)
        if bg: cell.fill=PatternFill("solid",fgColor=bg)
        if brd:
            t=Side(style='thin')
            cell.border=Border(left=t,right=t,top=t,bottom=t)

    ct_counter=[700]

    for jinfo in jornadas_activas:
        j_num         = jinfo['jornada_num']
        j_nombre_hoja = jinfo['jornada_nombre']
        fecha_inicio  = jinfo.get('fecha')

        df_jor = df_plan[df_plan['jornada']==j_nombre_hoja].copy()
        if len(df_jor)==0: continue

        ws = wb.create_sheet(title=f"J{j_num} — {j_nombre_hoja[:12]}")
        ws.sheet_view.showGridLines=False

        anchos={'A':14,'B':14,'C':8,'D':5,'E':5,'F':6,'G':6,'H':6,'I':5,
                'J':18,'K':13,'L':10,'M':14,'N':5}
        for col_l,w in anchos.items():
            ws.column_dimensions[col_l].width=w
        for i in range(dias_op):
            ws.column_dimensions[get_column_letter(15+i)].width=7
        ws.column_dimensions[get_column_letter(15+dias_op)].width=6

        cur=1
        equipos_jor=[e['nombre'] for e in eq_cfg if e['nombre'] in df_jor['equipo'].values]

        for grupo_num,nombre_eq in enumerate(equipos_jor,1):
            df_eq=df_jor[df_jor['equipo']==nombre_eq].copy()
            if len(df_eq)==0: continue

            pi=personal_info.get(nombre_eq,{})
            n_enc=next((e['enc'] for e in eq_cfg if e['nombre']==nombre_eq),3)
            last_col=15+dias_op

            if fecha_inicio:
                fechas=[fecha_inicio+timedelta(days=i) for i in range(dias_op)]
                fi_str=fecha_inicio.strftime("%d-%b-%y").upper()
                ff_str=fechas[-1].strftime("%d-%b-%y").upper()
            else:
                fechas=None; fi_str="____"; ff_str="____"

            def merge_row(row,c1,c2,val,**kw):
                ws.merge_cells(f'{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}')
                c=ws.cell(row,c1,val); sc(c,**kw); return c

            for txt in ["INSTITUTO NACIONAL DE ESTADÍSTICA Y CENSOS",
                        "COORDINACIÓN ZONAL LITORAL CZ8L",
                        "ACTUALIZACIÓN CARTOGRÁFICA - ENLISTAMIENTO",
                        "PROGRAMACIÓN OPERATIVO DE CAMPO"]:
                merge_row(cur,1,last_col,txt,bold=True,bg=AZ_OSCURO,fg=BLANCO,ha="center",sz=9)
                cur+=1
            cur+=1

            ws.cell(cur,1,"JORNADA"); sc(ws.cell(cur,1),bold=True,sz=10)
            ws.cell(cur,2,str(j_num)).font=Font(bold=True,size=11)
            ws.cell(cur,7,"GRUPO"); sc(ws.cell(cur,7),bold=True,sz=10)
            ws.cell(cur,9,str(grupo_num)).font=Font(bold=True,size=11)
            cur+=2

            ws.cell(cur,1,"PERÍODO DE ACTUALIZACIÓN:"); sc(ws.cell(cur,1),bold=True,sz=9)
            ws.cell(cur,5,"DEL"); ws.cell(cur,6,fi_str)
            ws.cell(cur,9,"AL");  ws.cell(cur,10,ff_str)
            cur+=2

            for col,txt in [(3,"COD."),(4,"NOMBRE"),(8,"No. CÉDULA"),(11,"No. CELULAR")]:
                sc(ws.cell(cur,col,txt),bold=True,sz=8)
            cur+=1

            ws.cell(cur,1,"SUPERVISOR:"); sc(ws.cell(cur,1),bold=True,sz=9)
            ws.cell(cur,3,pi.get('supervisor_cod',''))
            ws.cell(cur,4,pi.get('supervisor_nombre',''))
            ws.cell(cur,8,pi.get('supervisor_cedula',''))
            ws.cell(cur,11,pi.get('supervisor_celular',''))
            cur+=2

            enc_list=pi.get('encuestadores',[])
            for j in range(n_enc):
                info=enc_list[j] if j<len(enc_list) else {}
                ws.cell(cur,1,"ENCUESTADOR"); sc(ws.cell(cur,1),bold=True,sz=9)
                ws.cell(cur,3,info.get('cod','')); ws.cell(cur,4,info.get('nombre',''))
                ws.cell(cur,8,info.get('cedula','')); ws.cell(cur,11,info.get('celular',''))
                cur+=1
            cur+=1

            ws.cell(cur,1,"VEHÍCULO: CHOFER"); sc(ws.cell(cur,1),bold=True,sz=9)
            ws.cell(cur,4,pi.get('chofer_nombre','')); ws.cell(cur,8,pi.get('chofer_cedula',''))
            cur+=1
            ws.cell(cur,1,"PLACA:"); sc(ws.cell(cur,1),bold=True,sz=9)
            ws.cell(cur,4,pi.get('placa','')); cur+=2

            merge_row(cur,1,4,"EQUIPO",bold=True,bg=AZ_MEDIO,fg=BLANCO,ha="center",sz=8,brd=True)
            merge_row(cur,5,14,"IDENTIFICACIÓN",bold=True,bg=AZ_MEDIO,fg=BLANCO,ha="center",sz=8,brd=True)
            merge_row(cur,15,14+dias_op,"RECORRIDO DE LOS SECTORES — FECHA",
                      bold=True,bg=AZ_MEDIO,fg=BLANCO,ha="center",sz=7,brd=True)
            sc(ws.cell(cur,last_col,"# VIV"),bold=True,bg=AZ_MEDIO,fg=BLANCO,ha="center",sz=8,brd=True)
            ws.row_dimensions[cur].height=24; cur+=1

            for ci,h in enumerate(["SUPERVISOR","ENCUESTADOR","CARGA","PROV","CANTON",
                                    "CIUDAD/PARROQ","ZONA","SECTOR","MAN","CÓDIGO",
                                    "PROVINCIA","CANTÓN","CIUDAD","NRO EDIF"],1):
                sc(ws.cell(cur,ci,h),bold=True,bg=AZ_CLARO,ha="center",sz=7,brd=True,wrap=True)
            for i in range(dias_op):
                lbl=fechas[i].strftime("%d/%m") if fechas else f"D{i+1}"
                sc(ws.cell(cur,15+i,lbl),bold=True,bg=AZ_CLARO,ha="center",sz=7,brd=True)
            sc(ws.cell(cur,last_col,"# VIV"),bold=True,bg=AZ_CLARO,ha="center",sz=7,brd=True)
            ws.row_dimensions[cur].height=32; cur+=1

            df_sorted=df_eq.sort_values(['encuestador','dia_inicio']).copy()
            enc_actual=None; fila_enc=0; viv_enc_acum=0; enc_ci=-1

            for _,(_, rd) in enumerate(df_sorted.iterrows()):
                enc_id=int(rd.get('encuestador',0))
                if enc_id!=enc_actual and enc_actual is not None:
                    pal=ENC_PALETAS[enc_ci%len(ENC_PALETAS)]
                    enc_info=enc_list[enc_actual-1] if 0<enc_actual<=len(enc_list) else {}
                    merge_row(cur,1,9,f"SUBTOTAL {enc_info.get('nombre',f'Enc {enc_actual}')}",
                              bold=True,bg=pal["subtot"],fg=pal["hdr"],ha="right",sz=8)
                    for ci in range(10,last_col): sc(ws.cell(cur,ci,""),bg=pal["subtot"],brd=True)
                    sc(ws.cell(cur,last_col,viv_enc_acum),bold=True,ha="center",sz=9,
                       bg=pal["subtot"],fg=pal["hdr"],brd=True)
                    ws.row_dimensions[cur].height=14; cur+=1; viv_enc_acum=0
                if enc_id!=enc_actual:
                    enc_actual=enc_id; fila_enc=0; enc_ci=(enc_ci+1)%len(ENC_PALETAS)

                pal=ENC_PALETAS[enc_ci%len(ENC_PALETAS)]
                bg_row=pal["par"] if fila_enc%2==0 else pal["impar"]
                fila_enc+=1; viv_enc_acum+=int(rd.get('viv',0))

                p_cod=parse_codigo(str(rd['id_entidad']))
                enc_i=enc_list[enc_id-1] if 0<enc_id<=len(enc_list) else {}
                ct_str=f"CT{ct_counter[0]:03d}"; ct_counter[0]+=1

                cod_parr = f"{p_cod['prov']}{p_cod['canton']}{p_cod['ciudad_parroq']}"
                geo = catalogo_lookup.get(cod_parr, {})
                vals = [
                    pi.get('supervisor_cedula', ''),
                    enc_i.get('cedula', ''),
                    ct_str,
                    p_cod['prov'], p_cod['canton'],
                    p_cod['ciudad_parroq'],
                    p_cod['zona'], p_cod['sector'], p_cod['man'],
                    str(rd['id_entidad']),
                    geo.get('provincia_nombre', ''),
                    geo.get('canton_nombre', ''),
                    geo.get('parroquia_nombre', ''),
                    geo.get('fcode', ''),
                ]

                for ci,val in enumerate(vals,1):
                    sc(ws.cell(cur,ci,val),bg=AZ_CLARO if ci==10 else bg_row,
                       ha="center",sz=8,brd=True)

                d_ini=int(rd.get('dia_inicio',rd.get('dia_operativo',1)))
                d_fin=int(rd.get('dia_fin',d_ini))
                for i in range(dias_op):
                    if d_ini<=i+1<=d_fin:
                        sc(ws.cell(cur,15+i,"✓"),bold=True,bg=VRD_CHECK,ha="center",sz=11,brd=True)
                    else:
                        sc(ws.cell(cur,15+i,""),bg=bg_row,ha="center",brd=True)
                sc(ws.cell(cur,last_col,int(rd.get('viv',0))),ha="center",sz=8,brd=True,bg=bg_row)
                cur+=1

            if enc_actual is not None:
                pal=ENC_PALETAS[enc_ci%len(ENC_PALETAS)]
                enc_info=enc_list[enc_actual-1] if 0<enc_actual<=len(enc_list) else {}
                merge_row(cur,1,9,f"SUBTOTAL {enc_info.get('nombre',f'Enc {enc_actual}')}",
                          bold=True,bg=pal["subtot"],fg=pal["hdr"],ha="right",sz=8)
                for ci in range(10,last_col): sc(ws.cell(cur,ci,""),bg=pal["subtot"],brd=True)
                sc(ws.cell(cur,last_col,viv_enc_acum),bold=True,ha="center",sz=9,
                   bg=pal["subtot"],fg=pal["hdr"],brd=True)
                ws.row_dimensions[cur].height=14; cur+=1

            sc(ws.cell(cur,last_col-1,"TOTAL"),bold=True,ha="right",sz=8,bg=AZ_CLARO,brd=True)
            sc(ws.cell(cur,last_col,int(df_eq['viv'].sum())),bold=True,ha="center",
               sz=8,bg=AZ_CLARO,brd=True)
            cur+=4

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════
_defs = {
    "data_raw":None,"data_mes":None,"graph_G":None,
    "resultados_generados":False,"df_plan":None,
    "tsp_results":{},"road_paths":{},"resumen_bal":None,
    "sil_score":None,"n_bombero":0,
    "personal_info":{},
    "balance_log":[],"cv_ini_bal":None,"cv_fin_bal":None,
    "viv_por_cluster_antes":None,"viv_por_cluster_despues":None,
    "mes_inicio_cal": 7,
    "config_jornadas": {},
    "catalogo_df": None, "catalogo_lookup": {}, "catalogo_cols": {},
    # v5.1 — edición manual
    "edits_log": [],           # historial de cambios manuales
    "edit_counter": 0,         # contador de ediciones aplicadas
    "params":{
        "dias_op":12,"viv_min":40,"viv_max":70,"factor_r":1.8,
        "usar_bomb":False,"usar_gye":True,"dias_gye":3,"umbral_gye":5,
        "cv_objetivo":10,"max_iter_bal":300,"k_vecinos":8,
    },
    "equipos_cfg":[
        {"id":1,"nombre":"Equipo 1","enc":3},
        {"id":2,"nombre":"Equipo 2","enc":3},
        {"id":3,"nombre":"Equipo 3","enc":3},
    ],
}
for k,v in _defs.items():
    if k not in st.session_state: st.session_state[k]=v

# ══════════════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("### 🗺️ Encuesta Nacional")
    st.markdown("<p style='font-size:10px;color:#445566;margin-top:-8px'>INEC · Zonal Litoral</p>",
                unsafe_allow_html=True)
    st.divider()

    # PASO 1 — GeoPackage
    st.markdown("<div class='step'>PASO 1</div>",unsafe_allow_html=True)
    st.markdown("**Muestra (.gpkg)**")
    gpkg_f=st.file_uploader("GeoPackage",type=["gpkg"],key="gpkg_up")
    if gpkg_f:
        dissolve=st.radio("Nivel",["Por UPM","Por manzana"],index=0)
        if st.button("⚡ Procesar",use_container_width=True,type="primary"):
            with st.spinner("Leyendo geometrías..."):
                try:
                    with tempfile.NamedTemporaryFile(delete=False,suffix=".gpkg") as tmp:
                        tmp.write(gpkg_f.read()); p_tmp=tmp.name
                    data=cargar_gpkg(p_tmp,dissolve_upm=dissolve.startswith("Por UPM"))
                    os.unlink(p_tmp)
                    st.session_state.data_raw=data
                    st.session_state.resultados_generados=False
                    st.success(f"✓ {len(data):,} entidades")
                except Exception as e: st.error(str(e))
        if st.session_state.data_raw is not None:
            st.markdown("<span class='pill-ok'>✓ Listo</span>",unsafe_allow_html=True)
    else:
        st.markdown("<span class='pill-w'>⏳ Sin archivo</span>",unsafe_allow_html=True)

    st.divider()

    # PASO 2 — GraphML
    st.markdown("<div class='step'>PASO 2</div>",unsafe_allow_html=True)
    st.markdown("**Red vial (.graphml)**")
    gml_f=st.file_uploader("GraphML",type=["graphml"],key="gml_up")
    if gml_f:
        if st.button("⚡ Cargar grafo",use_container_width=True):
            with st.spinner("Cargando..."):
                try:
                    with tempfile.NamedTemporaryFile(delete=False,suffix=".graphml") as tmp:
                        tmp.write(gml_f.read()); pg=tmp.name
                    G=ox.load_graphml(pg); os.unlink(pg)
                    st.session_state.graph_G=G
                    st.success(f"✓ {len(G.nodes):,} nodos")
                except Exception as e: st.error(str(e))
        if st.session_state.graph_G is not None:
            st.markdown("<span class='pill-ok'>✓ Red lista</span>",unsafe_allow_html=True)
    else:
        st.markdown("<span class='pill-w'>⏳ Sin grafo</span>",unsafe_allow_html=True)

    st.divider()

    if st.session_state.data_raw is not None:
        # PASO 3 — Mes operativo
        st.markdown("<div class='step'>PASO 3</div>",unsafe_allow_html=True)
        st.markdown("**Mes operativo**")
        meses_disp=sorted(st.session_state.data_raw["mes"].dropna().unique().tolist())
        mes_sel=st.selectbox("Mes operativo",meses_disp,
            format_func=lambda x: f"Mes {int(x)}")
        df_mes=st.session_state.data_raw[st.session_state.data_raw["mes"]==mes_sel].copy()
        st.session_state.data_mes=df_mes

        st.markdown("**Mes calendario de inicio del operativo**")
        mes_ini_cal=st.selectbox(
            "El mes operativo 1 corresponde a:",
            list(MESES_CAL.keys()),
            index=st.session_state.mes_inicio_cal-1,
            format_func=lambda x: MESES_CAL[x],
            key="sel_mes_ini_cal"
        )
        st.session_state.mes_inicio_cal=mes_ini_cal

        j1_n,j2_n,mes_nom=jornada_num_desde_mes(int(mes_sel),mes_ini_cal)
        st.markdown(f"""
        <div style='font-size:11px;background:#0d2035;border-radius:6px;
                    padding:8px 12px;border-left:3px solid #2e86de;margin-top:6px'>
        📅 Mes {int(mes_sel)} ({mes_nom}) →
        <b style='color:#2e86de'>Jornada {j1_n}</b> +
        <b style='color:#27ae60'>Jornada {j2_n}</b>
        </div>""",unsafe_allow_html=True)

        st.divider()

        # PASO 4 — Equipos
        st.markdown("<div class='step'>PASO 4</div>",unsafe_allow_html=True)
        st.markdown("**Equipos**")
        c1,c2=st.columns(2)
        with c1:
            if st.button("＋",use_container_width=True):
                nid=max(t["id"] for t in st.session_state.equipos_cfg)+1
                st.session_state.equipos_cfg.append({"id":nid,"nombre":f"Equipo {nid}","enc":3})
                st.session_state.resultados_generados=False
        with c2:
            if st.button("－",use_container_width=True,
                         disabled=len(st.session_state.equipos_cfg)<=1):
                st.session_state.equipos_cfg.pop()
                st.session_state.resultados_generados=False

        for i,eq in enumerate(st.session_state.equipos_cfg):
            cc1,cc2=st.columns([2,1])
            with cc1:
                nn=st.text_input(f"n{eq['id']}",value=eq["nombre"],
                                 key=f"n_{eq['id']}",label_visibility="collapsed")
                st.session_state.equipos_cfg[i]["nombre"]=nn
            with cc2:
                ne=st.number_input("e",min_value=1,max_value=6,value=eq["enc"],
                                   key=f"e_{eq['id']}",label_visibility="collapsed")
                st.session_state.equipos_cfg[i]["enc"]=ne

        st.divider()
        st.markdown("**Parámetros operativos**")
        p=st.session_state.params

        # ── SLIDERS CON TOOLTIPS (v5.1) ──────────────────────────────────────
        p["dias_op"] = st.slider(
            "Días operativos", 10, 14, p["dias_op"],
            help="Cantidad de días laborables que dura cada jornada. "
                 "Cada jornada (quincenal) se divide en este número de días "
                 "para programar las visitas de campo. Valores típicos: 10-12."
        )
        p["viv_min"] = st.slider(
            "Mín viv/día", 30, 60, p["viv_min"],
            help="Número mínimo de viviendas que un encuestador debe cubrir "
                 "en un día operativo. Si una manzana tiene menos viviendas "
                 "que este valor, se combinará con otras manzanas en el mismo día. "
                 "Ajusta según la densidad urbana de la zona."
        )
        p["viv_max"] = st.slider(
            "Máx viv/día", 60, 120, p["viv_max"],
            help="Número máximo de viviendas que un encuestador puede cubrir "
                 "en un día operativo. Si una manzana supera este valor, se "
                 "le asignarán múltiples días consecutivos. Valores altos = "
                 "jornadas más largas para el encuestador."
        )
        p["factor_r"] = st.slider(
            "Factor rural (×)", 1.0, 2.5, p["factor_r"], 0.1,
            help="Multiplicador de carga para UPMs dispersas/rurales (tipo sector). "
                 "Un factor de 1.8× significa que una vivienda en zona dispersa "
                 "'pesa' 1.8 veces más que una urbana al distribuir la carga, "
                 "compensando el mayor tiempo de traslado entre viviendas rurales."
        )

        st.markdown("**Rebalanceo clusters**")
        p["cv_objetivo"] = st.slider(
            "CV objetivo (%)", 3, 25, p.get("cv_objetivo", 10),
            help="Coeficiente de variación objetivo entre clusters. Mide qué tan "
                 "equitativamente se distribuye la carga entre equipos. "
                 "Un CV bajo (3-5%) = distribución muy igualitaria pero puede "
                 "romper la contigüidad geográfica. Un CV alto (15-25%) = más "
                 "tolerancia a diferencias pero mejor cohesión territorial. "
                 "Recomendado: 8-12%."
        )
        p["max_iter_bal"] = st.slider(
            "Iter. máx.", 50, 500, p.get("max_iter_bal", 300), step=50,
            help="Número máximo de iteraciones del algoritmo de rebalanceo. "
                 "Más iteraciones = mejor balance posible pero mayor tiempo "
                 "de cómputo. El algoritmo se detendrá antes si alcanza el "
                 "CV objetivo o detecta un plateau (sin mejora significativa)."
        )
        p["k_vecinos"] = st.slider(
            "Vecinos frontera (k)", 4, 20, p.get("k_vecinos", 8),
            help="Número de vecinos geográficos más cercanos que se consideran "
                 "para el modo de swap 'frontera'. Un k más alto permite "
                 "intercambios entre clusters menos adyacentes, pero puede "
                 "reducir la compacidad geográfica. Recomendado: 6-10."
        )

        p["usar_bomb"] = st.toggle("Equipo Bombero", value=p["usar_bomb"],
            help="Activa la detección automática de UPMs outlier (muy lejanas "
                 "del centroide de su cluster) que se asignan a un 'Equipo Bombero' "
                 "especial. Útil cuando hay manzanas aisladas que desequilibrarían "
                 "la planificación de un equipo regular."
        )
        if p["usar_bomb"]:
            p["min_dist_bomb_m"] = st.slider(
                "Dist. mín. Bombero (km)", 10, 150,
                p.get("min_dist_bomb_m", 40000)//1000,
                help="Distancia mínima (en km) desde el centroide del cluster "
                     "para que una UPM sea candidata a Equipo Bombero. Solo las "
                     "UPMs que superen esta distancia Y sean outliers estadísticos "
                     "(IQR×3) serán reasignadas."
            )*1000

        p["usar_gye"] = st.toggle("Restricción Guayaquil", value=p["usar_gye"],
            help="Reserva los primeros días de la Jornada 1 exclusivamente para "
                 "UPMs dentro del cantón Guayaquil. Esto permite que los equipos "
                 "trabajen primero en la zona urbana densa antes de desplazarse "
                 "a zonas periféricas o rurales."
        )
        p["dias_gye"] = st.slider(
            "Días GYE", 1, 5, p["dias_gye"],
            disabled=not p["usar_gye"],
            help="Número de días al inicio de la Jornada 1 reservados exclusivamente "
                 "para trabajar dentro de Guayaquil. Los demás días se usan para "
                 "las UPMs fuera de GYE. Recomendado: 2-3 días."
        )
        p["umbral_gye"] = st.slider(
            "Umbral GYE (%)", 5, 30, p["umbral_gye"],
            disabled=not p["usar_gye"],
            help="Porcentaje mínimo de UPMs que deben pertenecer a Guayaquil "
                 "para que se active la restricción GYE. Si las UPMs de GYE son "
                 "menos que este porcentaje del total, la restricción se desactiva "
                 "automáticamente para ese mes."
        )

        tot_enc=sum(e["enc"] for e in st.session_state.equipos_cfg)
        tot_viv=int(df_mes["viv"].sum()) if len(df_mes)>0 else 0
        st.markdown(f"""
        <div style='font-size:11px;color:#445566;line-height:2;margin-top:8px'>
        📍 <b style='color:#7eb3d8'>{len(df_mes):,}</b> UPMs · mes {int(mes_sel)}<br>
        🏠 <b style='color:#7eb3d8'>{tot_viv:,}</b> viviendas<br>
        👥 <b style='color:#7eb3d8'>{len(st.session_state.equipos_cfg)}</b> equipos ·
           <b style='color:#7eb3d8'>{tot_enc}</b> enc.
        </div>""",unsafe_allow_html=True)

# ── HEADER ────────────────────────────────────
st.markdown("""
<div class='hdr'>
  <h1>Planificación Automática · Actualización Cartográfica</h1>
  <p>Encuesta Nacional &nbsp;·&nbsp; Zonal Litoral &nbsp;·&nbsp; INEC Ecuador</p>
</div>""",unsafe_allow_html=True)

if st.session_state.data_raw is None:
    st.markdown("<div class='ibox'>👈 Carga el <code>.gpkg</code> desde el panel lateral.</div>",
                unsafe_allow_html=True)
    st.stop()

df=st.session_state.data_mes
if df is None or len(df)==0:
    st.warning("Sin datos para el mes seleccionado."); st.stop()

p=st.session_state.params
k1,k2,k3,k4,k5=st.columns(5)
cv_v=cv_pct(df["viv"]); cv_c="#27ae60" if cv_v<50 else "#e74c3c"
for col,(val,lbl,sub,c) in zip([k1,k2,k3,k4,k5],[
    (f"{len(df):,}","UPMs",f"mes {int(df['mes'].iloc[0])}","#2e86de"),
    (f"{int(df['viv'].sum()):,}","Viviendas","precenso 2020","#2e86de"),
    (f"{len(df[df['tipo_entidad'].isin(['man','man_upm'])]):,}","Amanzanadas","man/man_upm","#2e86de"),
    (f"{len(df[df['tipo_entidad'].isin(['sec','sec_upm'])]):,}","Dispersas","sec/sec_upm","#2e86de"),
    (f"{cv_v:.1f}%","CV viviendas","dispersión",cv_c),
]):
    with col:
        st.markdown(f"<div class='kcard'><div class='v' style='color:{c}'>{val}</div>"
                    f"<div class='l'>{lbl}</div><div class='s'>{sub}</div></div>",
                    unsafe_allow_html=True)

st.markdown("<br>",unsafe_allow_html=True)

cb1,cb2=st.columns([1,3])
with cb1:
    btn=st.button("⚡ Generar Planificación",use_container_width=True,
                  type="primary",disabled=(st.session_state.graph_G is None))
with cb2:
    if st.session_state.graph_G is None:
        st.markdown("<div class='wbox'>⚠️ Carga el <code>.graphml</code> (Paso 2).</div>",
                    unsafe_allow_html=True)
    elif st.session_state.resultados_generados:
        st.markdown("<div class='ibox'>✓ Planificación lista. Puedes regenerar o editar manualmente.</div>",
                    unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
#  ALGORITMO PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════
if btn:
    G=st.session_state.graph_G
    eq_cfg=st.session_state.equipos_cfg
    n_eq=len(eq_cfg); nombres=[e["nombre"] for e in eq_cfg]
    n_clust=n_eq*2; p=st.session_state.params

    df_w=df.copy()
    df_w['equipo']='sin_asignar'; df_w['jornada']='sin_asignar'
    df_w['cluster_geo']=-1
    df_w['carga_pond']=df_w.apply(
        lambda r: r['viv']*p["factor_r"]
        if str(r.get('tipo_entidad','')).startswith('sec') else r['viv'],axis=1)
    df_w['encuestador']=0; df_w['dia_operativo']=0
    df_w['dia_inicio']=0;  df_w['dia_fin']=0; df_w['dist_base_m']=0.0

    prog=st.progress(0,"Iniciando...")

    prog.progress(5,"Distancias a base...")
    t_utm=Transformer.from_crs("EPSG:4326","EPSG:32717",always_xy=True)
    bx,by=t_utm.transform(BASE_LON,BASE_LAT)
    df_w['dist_base_m']=np.sqrt((df_w['x']-bx)**2+(df_w['y']-by)**2)

    prog.progress(8,"Verificando restricción Guayaquil...")
    upms_gye=pd.Series(False,index=df_w.index)
    if p["usar_gye"] and 'pro_x' in df_w.columns and 'can_x' in df_w.columns:
        upms_gye=(df_w['pro_x']==PRO_GYE)&(df_w['can_x']==CAN_GYE)
    pct_gye=upms_gye.sum()/len(df_w) if len(df_w)>0 else 0
    act_gye=p["usar_gye"] and (pct_gye>=p["umbral_gye"]/100) and upms_gye.sum()>0

    df_gye   =df_w[upms_gye].copy()  if act_gye else pd.DataFrame()
    df_no_gye=df_w[~upms_gye].copy()

    prog.progress(12,f"KMeans + rebalanceo ({n_clust} clusters)...")
    mask_bomb=pd.Series(False,index=df_w.index)

    if len(df_no_gye)>=n_clust:
        km_init=KMeans(n_clusters=n_clust,init='k-means++',n_init=20,
                       max_iter=500,random_state=42)
        lab_init=km_init.fit_predict(df_no_gye[['x','y']].values.astype(float))
        carg_arr=df_no_gye['carga_pond'].values
        vib_antes={c:float(carg_arr[lab_init==c].sum()) for c in range(n_clust)}
        st.session_state.viv_por_cluster_antes=vib_antes

        prog.progress(22,f"Rebalanceo (CV objetivo {p['cv_objetivo']}%)...")
        labels,bal_log,cv_ini,cv_fin=clustering_balanceado(
            df_no_gye,n_clusters=n_clust,
            cv_objetivo=p["cv_objetivo"]/100.0,
            max_iter=p["max_iter_bal"],k_vecinos=p["k_vecinos"])

        df_no_gye=df_no_gye.copy()
        df_no_gye['cluster_geo']=labels
        st.session_state.balance_log=bal_log
        st.session_state.cv_ini_bal=cv_ini
        st.session_state.cv_fin_bal=cv_fin

        try: st.session_state.sil_score=silhouette_score(df_no_gye[['x','y']].values,labels)
        except: st.session_state.sil_score=None

        vib_despues={c:float(carg_arr[labels==c].sum()) for c in range(n_clust)}
        st.session_state.viv_por_cluster_despues=vib_despues

        centroides=np.array([
            df_no_gye[['x','y']].values[labels==c].mean(axis=0)
            if (labels==c).sum()>0 else np.array([bx,by])
            for c in range(n_clust)])
        dist_c=np.sqrt((centroides[:,0]-bx)**2+(centroides[:,1]-by)**2)
        orden=np.argsort(dist_c)[::-1]
        asig={}
        for i,(cj1,cj2) in enumerate(zip(orden[:n_eq],orden[n_eq:])):
            asig[cj1]=(nombres[i],'Jornada 1')
            asig[cj2]=(nombres[i],'Jornada 2')

        df_no_gye['equipo'] =df_no_gye['cluster_geo'].map(lambda c: asig[c][0])
        df_no_gye['jornada']=df_no_gye['cluster_geo'].map(lambda c: asig[c][1])

        if p["usar_bomb"]:
            prog.progress(32,"Detectando outliers (Equipo Bombero)...")
            MIN_D=p.get("min_dist_bomb_m",40000)
            for c_id in range(n_clust):
                if c_id not in asig: continue
                mask_c=df_no_gye['cluster_geo']==c_id
                pts=df_no_gye[mask_c]
                if len(pts)<8: continue
                cx,cy=centroides[c_id]
                dists=np.sqrt((pts['x']-cx)**2+(pts['y']-cy)**2)
                Q1c,Q3c=dists.quantile(.25),dists.quantile(.75)
                iqrc=Q3c-Q1c
                if iqrc==0: continue
                bomb_idx=dists[(dists>Q3c+3*iqrc)&(dists>MIN_D)].index
                if len(bomb_idx)>0:
                    df_no_gye.loc[bomb_idx,'equipo']='Equipo Bombero'
                    df_no_gye.loc[bomb_idx,'jornada']='Jornada Especial'
                    mask_bomb.loc[bomb_idx]=True

        df_w.update(df_no_gye[['equipo','jornada','cluster_geo']])

    st.session_state.n_bombero=int((df_w['equipo']=='Equipo Bombero').sum())

    prog.progress(42,"Asignando encuestadores y días...")
    enc_dict={e["nombre"]:e["enc"] for e in eq_cfg}

    for nombre_eq in nombres:
        for jornada in ['Jornada 1','Jornada 2']:
            mask_g=(df_w['equipo']==nombre_eq)&(df_w['jornada']==jornada)
            grp=df_w[mask_g].copy()
            if len(grp)==0: continue
            n_enc=enc_dict.get(nombre_eq,3)

            if jornada=='Jornada 1' and act_gye:
                inicio=p["dias_gye"]+1; dias_disp=p["dias_op"]-p["dias_gye"]
            else:
                inicio=1; dias_disp=p["dias_op"]

            if dias_disp<=0: continue

            ga=asignar_encuestadores_y_dias(grp,n_enc,dias_disp,
                                            p["viv_min"],p["viv_max"],inicio)
            df_w.update(ga[['encuestador','dia_operativo','dia_inicio','dia_fin']])

    if act_gye and len(df_gye)>0:
        n_gye_clusters = min(n_eq, len(df_gye))
        if n_gye_clusters >= 2 and len(df_gye) >= n_gye_clusters:
            labels_gye, _, _, _ = clustering_balanceado(
                df_gye, n_clusters=n_gye_clusters,
                cv_objetivo=0.15, max_iter=200, k_vecinos=6)
            df_gye['cluster_gye'] = labels_gye

            for c_id in range(n_gye_clusters):
                eq_a = nombres[c_id % n_eq]
                mask_c = df_gye['cluster_gye'] == c_id
                grp_gye = df_gye[mask_c].copy()
                if len(grp_gye) == 0:
                    continue

                n_enc_eq = enc_dict.get(eq_a, 3)
                dias_gye = p["dias_gye"]

                grp_gye['equipo'] = eq_a
                grp_gye['jornada'] = 'Jornada 1'

                ga_gye = asignar_encuestadores_y_dias(
                    grp_gye, n_enc_eq, dias_gye,
                    p["viv_min"], p["viv_max"], inicio_dia=1)
                df_w.update(ga_gye[['equipo', 'jornada', 'encuestador',
                                     'dia_operativo', 'dia_inicio', 'dia_fin']])
        else:
            eq_a = nombres[0]
            df_gye['equipo'] = eq_a
            df_gye['jornada'] = 'Jornada 1'
            n_enc_eq = enc_dict.get(eq_a, 3)
            ga_gye = asignar_encuestadores_y_dias(
                df_gye, n_enc_eq, p["dias_gye"],
                p["viv_min"], p["viv_max"], inicio_dia=1)
            df_w.update(ga_gye[['equipo', 'jornada', 'encuestador',
                                 'dia_operativo', 'dia_inicio', 'dia_fin']])

    # TSP
    prog.progress(52,"Optimizando rutas TSP...")
    base_nd=ox.nearest_nodes(G,BASE_LON,BASE_LAT)
    G_u=G.to_undirected(); comp_base=nx.node_connected_component(G_u,base_nd)
    tsp_r,road_p={},{}

    for ri,nombre_eq in enumerate(nombres):
        for jornada in ['Jornada 1','Jornada 2']:
            pct=52+int((ri*2+['Jornada 1','Jornada 2'].index(jornada)+1)/(n_eq*2)*42)
            prog.progress(pct,f"TSP: {nombre_eq} | {jornada}...")
            mask_g=(df_w['equipo']==nombre_eq)&(df_w['jornada']==jornada)
            grp=df_w[mask_g]
            if len(grp)==0: continue
            nr=ox.nearest_nodes(G,grp['lon'].values,grp['lat'].values)
            nk=[n for n in nr if n in comp_base]
            if not nk: continue
            nu=[base_nd]+list(dict.fromkeys(nk)); n=len(nu)
            if n<=2: continue
            D=np.zeros((n,n))
            for i in range(n):
                for j in range(i+1,n):
                    try: d=nx.shortest_path_length(G_u,nu[i],nu[j],weight='length');D[i,j]=D[j,i]=d
                    except: D[i,j]=D[j,i]=1e9
            Gt=nx.Graph()
            for i in range(n):
                for j in range(i+1,n):
                    if D[i,j]<1e8: Gt.add_edge(i,j,weight=D[i,j])
            if not nx.is_connected(Gt): continue
            try: ciclo=approximation.traveling_salesman_problem(Gt,weight='weight',cycle=True)
            except: continue
            if 0 in ciclo:
                i0=ciclo.index(0); ciclo=ciclo[i0:]+ciclo[1:i0+1]
            dist=sum(D[ciclo[i],ciclo[i+1]] for i in range(len(ciclo)-1))
            ruta=[]; ng=[nu[idx] for idx in ciclo]
            for k in range(len(ng)-1):
                try:
                    seg=nx.shortest_path(G_u,ng[k],ng[k+1],weight='length')
                    ruta.extend((G.nodes[nd]['y'],G.nodes[nd]['x']) for nd in seg[:-1])
                except: continue
            if ng: ruta.append((G.nodes[ng[-1]]['y'],G.nodes[ng[-1]]['x']))
            clave=f"{nombre_eq}||{jornada}"
            tsp_r[clave]={'equipo':nombre_eq,'jornada':jornada,
                          'n_puntos':len(grp),'dist_km':dist/1000}
            road_p[clave]=ruta

    prog.progress(98,"Métricas finales...")
    resumen=df_w[~df_w['equipo'].isin(['Equipo Bombero','sin_asignar'])].groupby(
        ['equipo','jornada']).agg(
        n_upms=('id_entidad','count'),
        viv_reales=('viv','sum'),
        carga_ponderada=('carga_pond','sum')).reset_index()
    dist_df=pd.DataFrame([
        {'equipo':v['equipo'],'jornada':v['jornada'],'dist_km':round(v['dist_km'],1)}
        for v in tsp_r.values()
    ]) if tsp_r else pd.DataFrame(columns=['equipo','jornada','dist_km'])
    resumen_bal=pd.merge(resumen,dist_df,on=['equipo','jornada'],how='left').fillna(0)

    prog.progress(100,"¡Listo!"); prog.empty()
    st.session_state.df_plan=df_w
    st.session_state.tsp_results=tsp_r; st.session_state.road_paths=road_p
    st.session_state.resumen_bal=resumen_bal
    st.session_state.resultados_generados=True
    st.session_state.edits_log = []
    st.session_state.edit_counter = 0
    st.success("✓ Planificación v5.1 generada.")

# ── RESULTADOS ────────────────────────────────
if not st.session_state.resultados_generados:
    st.markdown("<div class='ibox'>👆 Presiona <b>Generar Planificación</b>.</div>",
                unsafe_allow_html=True)
    st.stop()

df_plan=st.session_state.df_plan
tsp_r=st.session_state.tsp_results; road_p=st.session_state.road_paths
res_bal=st.session_state.resumen_bal; eq_cfg=st.session_state.equipos_cfg
nombres=[e["nombre"] for e in eq_cfg]; p=st.session_state.params
j1_n,j2_n,mes_nom=jornada_num_desde_mes(
    int(df['mes'].iloc[0]), st.session_state.mes_inicio_cal)

color_map={n:COLORES[i%len(COLORES)] for i,n in enumerate(nombres)}
color_map['Equipo Bombero']='#9b59b6'

# ══════════════════════════════════════════════════════════════════════════════
#  TABS (v5.1 — incluye Edición Manual)
# ══════════════════════════════════════════════════════════════════════════════
tab_mapa, tab_analisis, tab_edicion, tab_plan, tab_reporte = st.tabs([
    "🗺️  Mapa de Rutas",
    "📊  Análisis de Carga",
    "✏️  Edición Manual",
    "📅  Planificación de Jornadas",
    "📋  Reporte y Descarga"
])

# ══ TAB 1 — MAPA ══════════════════════════════
with tab_mapa:
    st.markdown("<div class='stitle'>Mapa del Operativo de Campo</div>",unsafe_allow_html=True)
    cc1,cc2=st.columns([1,3])
    with cc1:
        mj1=st.checkbox("Jornada 1",value=True)
        mj2=st.checkbox("Jornada 2",value=True)
        n_b=int((df_plan['equipo']=='Equipo Bombero').sum())
        mbm=st.checkbox(f"Equipo Bombero ({n_b})",value=True)
        mrts=st.checkbox("Mostrar rutas",value=True)
        fnd=st.selectbox("Fondo",["CartoDB dark_matter","CartoDB positron","OpenStreetMap"])
        st.divider()
        st.markdown("**Leyenda:**")
        for n,c in color_map.items():
            if n in nombres:
                st.markdown(f"<span style='color:{c};font-size:17px'>●</span> {n}",
                            unsafe_allow_html=True)
        st.markdown(f"<span style='color:#9b59b6;font-size:17px'>●</span> Equipo Bombero ({n_b})",
                    unsafe_allow_html=True)
        # Indicador de ediciones manuales
        n_edits = st.session_state.edit_counter
        if n_edits > 0:
            st.markdown(f"<div class='edit-card'>✏️ {n_edits} edición(es) manual(es) aplicada(s)</div>",
                        unsafe_allow_html=True)
    with cc2:
        m=folium.Map(location=[BASE_LAT,BASE_LON],zoom_start=8,tiles=fnd)
        folium.Marker([BASE_LAT,BASE_LON],popup="<b>Base INEC GYE</b>",
            icon=folium.Icon(color='white',icon='home',prefix='fa')).add_to(m)
        for _,row in df_plan.iterrows():
            eq,jor=row.get('equipo',''),row.get('jornada','')
            if jor=='Jornada 1' and not mj1: continue
            if jor=='Jornada 2' and not mj2: continue
            if jor=='Jornada Especial' and not mbm: continue
            clr=color_map.get(eq,'#888')
            d_ini=int(row.get('dia_inicio',row.get('dia_operativo',0)))
            d_fin=int(row.get('dia_fin',d_ini))
            dias_str=f"Día {d_ini}" if d_ini==d_fin else f"Días {d_ini}–{d_fin}"
            folium.CircleMarker(
                location=[row['lat'],row['lon']],radius=5,color=clr,
                fill=True,fill_color=clr,fill_opacity=.85,
                popup=folium.Popup(
                    f"<b>ID:</b> {row['id_entidad']}<br><b>Viv:</b> {int(row['viv'])}<br>"
                    f"<b>Equipo:</b> {eq}<br><b>Jornada:</b> {jor}<br>"
                    f"<b>Enc:</b> {int(row.get('encuestador',0))}<br>"
                    f"<b>{dias_str}</b>",max_width=210),
                tooltip=f"{eq}·Enc{int(row.get('encuestador',0))}·{int(row['viv'])}viv"
            ).add_to(m)
        if mrts:
            for clave,coords in road_p.items():
                eq,jor=clave.split('||')
                if jor=='Jornada 1' and not mj1: continue
                if jor=='Jornada 2' and not mj2: continue
                if len(coords)>1:
                    folium.PolyLine(coords,weight=3,color=color_map.get(eq,'#888'),
                                    opacity=.75,tooltip=f"{eq}|{jor}").add_to(m)
        st_folium(m,width=None,height=540,returned_objects=[],key="mapa_v5")

# ══ TAB 2 — ANÁLISIS ══════════════════════════
with tab_analisis:
    st.markdown("<div class='stitle'>Rebalanceo de Clusters</div>",unsafe_allow_html=True)

    cv_ini=st.session_state.get("cv_ini_bal")
    cv_fin=st.session_state.get("cv_fin_bal")
    bal_log=st.session_state.get("balance_log",[])
    viv_ant=st.session_state.get("viv_por_cluster_antes")
    viv_dep=st.session_state.get("viv_por_cluster_despues")

    if cv_ini is not None and cv_fin is not None:
        mejora=cv_ini-cv_fin
        cc_m="#27ae60" if mejora>5 else ("#f39c12" if mejora>0 else "#e74c3c")
        modo_fin=next((l.get('modo','') for l in reversed(bal_log)
                      if 'objetivo' in l.get('modo','') or 'plateau' in l.get('modo','')
                      or 'sin mejora' in l.get('modo','')),'-')
        st.markdown(f"""
        <div class='balance-box'>
        <b>CV inicial (KMeans puro):</b>
        <span style='color:#e74c3c;font-family:monospace'>{cv_ini:.1f}%</span>
        &nbsp;→&nbsp;
        <b>CV final (rebalanceo):</b>
        <span style='color:#27ae60;font-family:monospace'>{cv_fin:.1f}%</span>
        &nbsp;&nbsp;<b style='color:{cc_m}'>Δ {mejora:.1f} pp</b><br>
        <span style='font-size:11px'>
        Parada: <i>{modo_fin}</i> ·
        Iteraciones: {len([l for l in bal_log if l.get('modo') not in ['inicial','final']])}
        </span>
        </div>""",unsafe_allow_html=True)

    if viv_ant and viv_dep:
        n_cl=len(viv_ant)
        df_comp=pd.DataFrame({
            'Cluster':[f"C{c}" for c in range(n_cl)]*2,
            'Carga pond.':(list(viv_ant.values())+list(viv_dep.values())),
            'Fase':['Antes (KMeans)']*n_cl+['Después (rebalanceo)']*n_cl
        })
        fig_comp=px.bar(df_comp,x='Cluster',y='Carga pond.',color='Fase',
                        barmode='group',title='Carga por cluster — antes vs después',
                        template='plotly_dark',
                        color_discrete_map={'Antes (KMeans)':'#e74c3c',
                                            'Después (rebalanceo)':'#27ae60'})
        fig_comp.update_layout(paper_bgcolor="#111827",plot_bgcolor="#0a1020",
                               title_font_size=12)
        st.plotly_chart(fig_comp,use_container_width=True)
        with st.expander("Historial de iteraciones"):
            if bal_log:
                st.dataframe(pd.DataFrame(bal_log),use_container_width=True,height=250)

    st.divider()
    st.markdown("<div class='stitle'>Equidad entre equipos</div>",unsafe_allow_html=True)
    df_main=df_plan[~df_plan['equipo'].isin(['Equipo Bombero','sin_asignar'])].copy()
    res_cv=df_main.groupby(['equipo','jornada']).agg(
        viv_reales=('viv','sum'),carga_ponderada=('carga_pond','sum')).reset_index()
    for jornada in ['Jornada 1','Jornada 2']:
        sub=res_cv[res_cv['jornada']==jornada]
        if len(sub)==0: continue
        if len(sub)==1:
            st.markdown(f"<div class='ibox'><b>{jornada}:</b> 1 equipo.</div>",
                        unsafe_allow_html=True); continue
        cr=cv_pct(sub['viv_reales']); cp=cv_pct(sub['carga_ponderada'])
        ccr="#27ae60" if cr<20 else ("#f39c12" if cr<40 else "#e74c3c")
        ccp="#27ae60" if cp<20 else ("#f39c12" if cp<40 else "#e74c3c")
        em="✓" if cp<20 else ("⚠" if cp<40 else "✗")
        st.markdown(f"""<div class='ibox'><b>{jornada}</b><br>
        &nbsp;&nbsp;CV viv. reales: <span style='color:{ccr};font-family:monospace;
        font-weight:600'>{cr:.1f}%</span><br>
        &nbsp;&nbsp;CV carga pond.: <span style='color:{ccp};font-family:monospace;
        font-weight:600'>{cp:.1f}%</span> {em}</div>""",unsafe_allow_html=True)

    with st.expander("Tabla de balance"): st.dataframe(res_cv,use_container_width=True)

    st.markdown("<div class='stitle'>Carga por equipo</div>",unsafe_allow_html=True)
    eq_act=[n for n in nombres if n in df_plan['equipo'].values]
    cols_e=st.columns(len(eq_act)) if eq_act else []
    for col_e,nombre_eq in zip(cols_e,eq_act):
        sub_e=df_plan[df_plan['equipo']==nombre_eq]
        vt=int(sub_e['viv'].sum()); cv_e=cv_pct(sub_e['carga_pond'])
        ce=color_map.get(nombre_eq,'#2e86de')
        ccv="#27ae60" if cv_e<20 else ("#f39c12" if cv_e<40 else "#e74c3c")
        with col_e:
            st.markdown(f"""<div class='eq-card' style='border-color:{ce}55'>
              <div style='width:10px;height:10px;background:{ce};border-radius:50%;margin:0 auto 7px'></div>
              <div style='font-family:"IBM Plex Mono",monospace;font-size:12px;
                          color:{ce};font-weight:600'>{nombre_eq}</div>
              <div style='font-size:17px;font-weight:600;color:#d0d8e8;margin:4px 0'>{vt:,}</div>
              <div style='font-size:10px;color:#7a8fa6'>viviendas</div>
              <div style='font-size:11px;color:{ccv};margin-top:4px'>CV {cv_e:.1f}%</div>
            </div>""",unsafe_allow_html=True)

    st.markdown("<br>",unsafe_allow_html=True)
    eq_sel=st.selectbox("Detalle:",eq_act)
    df_sel=df_plan[df_plan['equipo']==eq_sel].copy()
    df_enc=df_sel.groupby(['jornada','encuestador']).agg(
        upms=('id_entidad','count'),viv_reales=('viv','sum'),
        carga_pond=('carga_pond','sum')).reset_index()
    cd1,cd2=st.columns(2)
    with cd1:
        fig=px.bar(df_enc,x='encuestador',y='viv_reales',color='jornada',barmode='group',
                   title=f'Viviendas — {eq_sel}',template='plotly_dark',
                   color_discrete_sequence=['#2e86de','#27ae60'])
        fig.update_layout(paper_bgcolor="#111827",plot_bgcolor="#0a1020",title_font_size=12)
        st.plotly_chart(fig,use_container_width=True)
    with cd2:
        fig2=px.bar(df_enc,x='encuestador',y='carga_pond',color='jornada',barmode='group',
                    title=f'Carga pond. — {eq_sel}',template='plotly_dark',
                    color_discrete_sequence=['#e74c3c','#f39c12'])
        fig2.update_layout(paper_bgcolor="#111827",plot_bgcolor="#0a1020",title_font_size=12)
        st.plotly_chart(fig2,use_container_width=True)

    st.markdown("<div class='stitle'>Distribución diaria</div>",unsafe_allow_html=True)
    jor_opts=["Jornada 1","Jornada 2","Ambas"]
    jor_filtro=st.radio("Filtrar:",jor_opts,horizontal=True,key="radio_jor_v5",index=0)
    pivot_all=df_plan[df_plan['equipo'].isin(eq_act)].copy()
    if jor_filtro!="Ambas": pivot_all=pivot_all[pivot_all['jornada']==jor_filtro]
    rows_exp=[]
    for _,row in pivot_all.iterrows():
        d_ini=int(row.get('dia_inicio',row.get('dia_operativo',1)))
        d_fin=int(row.get('dia_fin',d_ini))
        dias_dur=max(1,d_fin-d_ini+1); viv_d=row['viv']/dias_dur
        for dd in range(d_ini,d_fin+1):
            rows_exp.append({'equipo':row['equipo'],'jornada':row['jornada'],
                             'encuestador':int(row.get('encuestador',0)),
                             'dia_abs':dd,'viv':viv_d})
    if rows_exp:
        df_exp=pd.DataFrame(rows_exp)
        if jor_filtro!="Ambas":
            df_exp['dia_rel']=df_exp['dia_abs']-df_exp['dia_abs'].min()+1
        else:
            df_exp['dia_rel']=df_exp.groupby('jornada')['dia_abs'].transform(
                lambda s: s-s.min()+1).astype(int)
        pivot=df_exp.groupby(['equipo','dia_rel'])['viv'].sum().reset_index()
        fig_d=px.bar(pivot,x='dia_rel',y='viv',color='equipo',barmode='group',
                     title=f'Viv/día — {jor_filtro}',
                     labels={'dia_rel':'Día','viv':'Viviendas'},
                     template='plotly_dark',color_discrete_map=color_map)
        tot_enc_f=sum(e["enc"] for e in eq_cfg if e["nombre"] in eq_act)
        avg_enc_f=tot_enc_f/max(1,len(eq_act))
        fig_d.add_hline(y=p["viv_min"]*avg_enc_f,line_dash="dot",line_color="#f39c12",
                        annotation_text=f"Mín ({p['viv_min']})")
        fig_d.add_hline(y=p["viv_max"]*avg_enc_f,line_dash="dot",line_color="#e74c3c",
                        annotation_text=f"Máx ({p['viv_max']})")
        fig_d.update_layout(paper_bgcolor="#111827",plot_bgcolor="#0a1020",
                            xaxis=dict(dtick=1))
        st.plotly_chart(fig_d,use_container_width=True)
        if jor_filtro!="Ambas":
            piv_enc=df_exp.groupby(['encuestador','dia_rel'])['viv'].sum().reset_index()
            piv_enc['encuestador']="Enc. "+piv_enc['encuestador'].astype(str)
            fig_enc=px.line(piv_enc,x='dia_rel',y='viv',color='encuestador',markers=True,
                            title=f'Carga diaria por encuestador — {jor_filtro}',
                            template='plotly_dark')
            fig_enc.add_hline(y=p["viv_min"],line_dash="dot",line_color="#f39c12")
            fig_enc.add_hline(y=p["viv_max"],line_dash="dot",line_color="#e74c3c")
            fig_enc.update_layout(paper_bgcolor="#111827",plot_bgcolor="#0a1020",
                                  xaxis=dict(dtick=1))
            st.plotly_chart(fig_enc,use_container_width=True)

    df_bm=df_plan[df_plan['equipo']=='Equipo Bombero']
    n_bm=st.session_state.n_bombero
    st.markdown("<div class='stitle'>Equipo Bombero</div>",unsafe_allow_html=True)
    if n_bm==0:
        st.markdown("<div class='bcard'><b style='color:#9b59b6'>Equipo Bombero</b> — 0 UPMs.</div>",
                    unsafe_allow_html=True)
    else:
        st.markdown(f"<div class='bcard'><b style='color:#9b59b6'>Equipo Bombero</b> — {n_bm} UPMs · "
                    f"{int(df_bm['viv'].sum()):,} viv</div>",unsafe_allow_html=True)
        st.dataframe(df_bm[['id_entidad','tipo_entidad','viv','lat','lon','dist_base_m']]
            .sort_values('dist_base_m',ascending=False).reset_index(drop=True),
            use_container_width=True,height=200)


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 3 — EDICIÓN MANUAL (NUEVO v5.1)
# ══════════════════════════════════════════════════════════════════════════════
with tab_edicion:
    st.markdown("<div class='stitle'>Edición Manual de la Planificación</div>",
                unsafe_allow_html=True)

    st.markdown("""<div class='ibox'>
    Aquí puedes ajustar manualmente la planificación automática: mover manzanas/UPMs
    a otro <b>equipo</b>, cambiar de <b>jornada</b>, reasignar <b>encuestador</b>
    o modificar los <b>días operativos</b>. Los cambios se aplican directamente
    sobre el plan y se reflejan en el mapa, los gráficos y el Excel final.
    </div>""", unsafe_allow_html=True)

    # Historial de ediciones
    n_edits = st.session_state.edit_counter
    if n_edits > 0:
        st.markdown(f"<div class='edit-ok'>"
                    f"<span class='edit-count'>{n_edits}</span> "
                    f"edición(es) manual(es) aplicadas en esta sesión</div>",
                    unsafe_allow_html=True)

    # ── Selector de modo de edición ──────────────────────────────────────────
    modo_edit = st.radio(
        "Modo de edición",
        ["Selección individual", "Selección por lote (filtro)"],
        horizontal=True,
        key="modo_edicion_radio",
        help="Individual: selecciona UPMs específicas por su código. "
             "Por lote: filtra por equipo/jornada y aplica cambios masivos."
    )

    # Opciones de destino (comunes a ambos modos)
    equipos_destino = nombres + ['Equipo Bombero']
    jornadas_destino = ['Jornada 1', 'Jornada 2', 'Jornada Especial']

    if modo_edit == "Selección individual":
        # ── MODO INDIVIDUAL ──────────────────────────────────────────────────
        st.markdown("#### Seleccionar UPMs para editar")

        # Filtros de contexto
        fc1, fc2 = st.columns(2)
        with fc1:
            ctx_equipo = st.selectbox(
                "Filtrar por equipo (contexto)",
                ["Todos"] + equipos_destino,
                key="ctx_eq_filter"
            )
        with fc2:
            ctx_jornada = st.selectbox(
                "Filtrar por jornada (contexto)",
                ["Todas"] + jornadas_destino,
                key="ctx_jor_filter"
            )

        # Aplicar filtro de contexto
        df_ctx = df_plan.copy()
        if ctx_equipo != "Todos":
            df_ctx = df_ctx[df_ctx['equipo'] == ctx_equipo]
        if ctx_jornada != "Todas":
            df_ctx = df_ctx[df_ctx['jornada'] == ctx_jornada]

        if len(df_ctx) == 0:
            st.warning("No hay UPMs con ese filtro.")
        else:
            # Mostrar tabla filtrada
            cols_show = ['id_entidad', 'viv', 'equipo', 'jornada',
                         'encuestador', 'dia_inicio', 'dia_fin']
            cols_show = [c for c in cols_show if c in df_ctx.columns]

            st.markdown(f"**{len(df_ctx):,} UPMs** en este filtro")

            # Selección de UPMs por multiselect
            ids_disponibles = df_ctx['id_entidad'].astype(str).tolist()
            ids_seleccionados = st.multiselect(
                "Selecciona UPMs a editar (puedes buscar por código):",
                ids_disponibles,
                key="edit_upm_select",
                help="Escribe parte del código para buscar. "
                     "Puedes seleccionar varias UPMs a la vez."
            )

            if ids_seleccionados:
                # Mostrar las UPMs seleccionadas
                mask_sel = df_plan['id_entidad'].astype(str).isin(ids_seleccionados)
                df_sel_edit = df_plan[mask_sel]
                st.dataframe(
                    df_sel_edit[cols_show].reset_index(drop=True),
                    use_container_width=True,
                    height=min(200, 35 * len(df_sel_edit) + 38)
                )

                st.markdown("---")
                st.markdown("#### Nuevos valores para las UPMs seleccionadas")
                st.markdown("<div class='wbox'>Deja en blanco los campos que no quieras cambiar.</div>",
                            unsafe_allow_html=True)

                ec1, ec2, ec3 = st.columns(3)
                with ec1:
                    nuevo_equipo = st.selectbox(
                        "Nuevo equipo",
                        ["— Sin cambio —"] + equipos_destino,
                        key="edit_nuevo_equipo"
                    )
                with ec2:
                    nueva_jornada = st.selectbox(
                        "Nueva jornada",
                        ["— Sin cambio —"] + jornadas_destino,
                        key="edit_nueva_jornada"
                    )
                with ec3:
                    max_enc = max(e['enc'] for e in eq_cfg)
                    nuevo_enc = st.selectbox(
                        "Nuevo encuestador",
                        ["— Sin cambio —"] + [str(i) for i in range(1, max_enc + 1)],
                        key="edit_nuevo_enc"
                    )

                ed1, ed2 = st.columns(2)
                with ed1:
                    nuevo_dia_ini = st.number_input(
                        "Nuevo día inicio (0 = sin cambio)",
                        min_value=0, max_value=p["dias_op"],
                        value=0, key="edit_dia_ini",
                        help="Día operativo de inicio (1 a días_op). "
                             "Deja en 0 para no cambiar."
                    )
                with ed2:
                    nuevo_dia_fin = st.number_input(
                        "Nuevo día fin (0 = sin cambio)",
                        min_value=0, max_value=p["dias_op"],
                        value=0, key="edit_dia_fin",
                        help="Día operativo de fin (1 a días_op). "
                             "Deja en 0 para no cambiar."
                    )

                # Resumen previo de cambios
                cambios = []
                if nuevo_equipo != "— Sin cambio —":
                    cambios.append(f"Equipo → **{nuevo_equipo}**")
                if nueva_jornada != "— Sin cambio —":
                    cambios.append(f"Jornada → **{nueva_jornada}**")
                if nuevo_enc != "— Sin cambio —":
                    cambios.append(f"Encuestador → **{nuevo_enc}**")
                if nuevo_dia_ini > 0:
                    cambios.append(f"Día inicio → **{nuevo_dia_ini}**")
                if nuevo_dia_fin > 0:
                    cambios.append(f"Día fin → **{nuevo_dia_fin}**")

                if cambios:
                    st.markdown(f"**Resumen:** {len(ids_seleccionados)} UPM(s) · "
                                + " · ".join(cambios))

                    if st.button("✅ Aplicar cambios", type="primary",
                                 key="btn_aplicar_individual",
                                 use_container_width=True):
                        idx_edit = df_plan.index[mask_sel]
                        log_entry = {
                            'tipo': 'individual',
                            'upms': ids_seleccionados.copy(),
                            'cambios': {}
                        }
                        if nuevo_equipo != "— Sin cambio —":
                            df_plan.loc[idx_edit, 'equipo'] = nuevo_equipo
                            log_entry['cambios']['equipo'] = nuevo_equipo
                        if nueva_jornada != "— Sin cambio —":
                            df_plan.loc[idx_edit, 'jornada'] = nueva_jornada
                            log_entry['cambios']['jornada'] = nueva_jornada
                        if nuevo_enc != "— Sin cambio —":
                            df_plan.loc[idx_edit, 'encuestador'] = int(nuevo_enc)
                            log_entry['cambios']['encuestador'] = int(nuevo_enc)
                        if nuevo_dia_ini > 0:
                            df_plan.loc[idx_edit, 'dia_inicio'] = nuevo_dia_ini
                            df_plan.loc[idx_edit, 'dia_operativo'] = nuevo_dia_ini
                            log_entry['cambios']['dia_inicio'] = nuevo_dia_ini
                        if nuevo_dia_fin > 0:
                            df_plan.loc[idx_edit, 'dia_fin'] = nuevo_dia_fin
                            log_entry['cambios']['dia_fin'] = nuevo_dia_fin

                        st.session_state.df_plan = df_plan
                        st.session_state.edits_log.append(log_entry)
                        st.session_state.edit_counter += 1

                        # Recalcular resumen
                        resumen = df_plan[~df_plan['equipo'].isin(
                            ['Equipo Bombero', 'sin_asignar'])].groupby(
                            ['equipo', 'jornada']).agg(
                            n_upms=('id_entidad', 'count'),
                            viv_reales=('viv', 'sum'),
                            carga_ponderada=('carga_pond', 'sum')).reset_index()
                        st.session_state.resumen_bal = resumen

                        st.success(f"✓ {len(ids_seleccionados)} UPM(s) actualizadas.")
                        st.rerun()
                else:
                    st.markdown("<div class='wbox'>Selecciona al menos un campo para cambiar.</div>",
                                unsafe_allow_html=True)

    else:
        # ── MODO POR LOTE ────────────────────────────────────────────────────
        st.markdown("#### Selección por lote (filtro masivo)")

        bl1, bl2, bl3 = st.columns(3)
        with bl1:
            lote_equipo_orig = st.selectbox(
                "Equipo origen",
                equipos_destino,
                key="lote_eq_orig"
            )
        with bl2:
            lote_jornada_orig = st.selectbox(
                "Jornada origen",
                jornadas_destino,
                key="lote_jor_orig"
            )
        with bl3:
            lote_enc_orig = st.selectbox(
                "Encuestador origen",
                ["Todos"] + [str(i) for i in range(1, max(e['enc'] for e in eq_cfg) + 1)],
                key="lote_enc_orig"
            )

        mask_lote = (
            (df_plan['equipo'] == lote_equipo_orig) &
            (df_plan['jornada'] == lote_jornada_orig)
        )
        if lote_enc_orig != "Todos":
            mask_lote = mask_lote & (df_plan['encuestador'] == int(lote_enc_orig))

        df_lote = df_plan[mask_lote]

        if len(df_lote) == 0:
            st.info("No hay UPMs con ese filtro.")
        else:
            st.markdown(f"**{len(df_lote):,} UPMs** coinciden con el filtro "
                        f"({int(df_lote['viv'].sum()):,} viviendas)")

            cols_show_l = ['id_entidad', 'viv', 'encuestador', 'dia_inicio', 'dia_fin']
            cols_show_l = [c for c in cols_show_l if c in df_lote.columns]
            with st.expander(f"Ver {len(df_lote)} UPMs del lote"):
                st.dataframe(
                    df_lote[cols_show_l].sort_values(
                        ['encuestador', 'dia_inicio']).reset_index(drop=True),
                    use_container_width=True,
                    height=300
                )

            st.markdown("---")
            st.markdown("#### Destino para todo el lote")

            ld1, ld2, ld3 = st.columns(3)
            with ld1:
                lote_equipo_dest = st.selectbox(
                    "Mover a equipo",
                    ["— Sin cambio —"] + equipos_destino,
                    key="lote_eq_dest"
                )
            with ld2:
                lote_jornada_dest = st.selectbox(
                    "Mover a jornada",
                    ["— Sin cambio —"] + jornadas_destino,
                    key="lote_jor_dest"
                )
            with ld3:
                lote_enc_dest = st.selectbox(
                    "Asignar a encuestador",
                    ["— Sin cambio —"] + [str(i) for i in range(1, max(e['enc'] for e in eq_cfg) + 1)],
                    key="lote_enc_dest"
                )

            cambios_lote = []
            if lote_equipo_dest != "— Sin cambio —":
                cambios_lote.append(f"Equipo → **{lote_equipo_dest}**")
            if lote_jornada_dest != "— Sin cambio —":
                cambios_lote.append(f"Jornada → **{lote_jornada_dest}**")
            if lote_enc_dest != "— Sin cambio —":
                cambios_lote.append(f"Encuestador → **{lote_enc_dest}**")

            if cambios_lote:
                st.markdown(f"**Cambio masivo:** {len(df_lote)} UPMs · "
                            + " · ".join(cambios_lote))

                if st.button("✅ Aplicar cambio masivo", type="primary",
                             key="btn_aplicar_lote",
                             use_container_width=True):
                    idx_lote = df_plan.index[mask_lote]
                    log_entry = {
                        'tipo': 'lote',
                        'origen': f"{lote_equipo_orig}/{lote_jornada_orig}/{lote_enc_orig}",
                        'n_upms': len(df_lote),
                        'cambios': {}
                    }
                    if lote_equipo_dest != "— Sin cambio —":
                        df_plan.loc[idx_lote, 'equipo'] = lote_equipo_dest
                        log_entry['cambios']['equipo'] = lote_equipo_dest
                    if lote_jornada_dest != "— Sin cambio —":
                        df_plan.loc[idx_lote, 'jornada'] = lote_jornada_dest
                        log_entry['cambios']['jornada'] = lote_jornada_dest
                    if lote_enc_dest != "— Sin cambio —":
                        df_plan.loc[idx_lote, 'encuestador'] = int(lote_enc_dest)
                        log_entry['cambios']['encuestador'] = int(lote_enc_dest)

                    st.session_state.df_plan = df_plan
                    st.session_state.edits_log.append(log_entry)
                    st.session_state.edit_counter += 1

                    resumen = df_plan[~df_plan['equipo'].isin(
                        ['Equipo Bombero', 'sin_asignar'])].groupby(
                        ['equipo', 'jornada']).agg(
                        n_upms=('id_entidad', 'count'),
                        viv_reales=('viv', 'sum'),
                        carga_ponderada=('carga_pond', 'sum')).reset_index()
                    st.session_state.resumen_bal = resumen

                    st.success(f"✓ {len(df_lote)} UPMs actualizadas en lote.")
                    st.rerun()
            else:
                st.markdown("<div class='wbox'>Selecciona al menos un cambio de destino.</div>",
                            unsafe_allow_html=True)

    # ── Historial de ediciones ────────────────────────────────────────────────
    st.markdown("<div class='stitle'>Historial de ediciones manuales</div>",
                unsafe_allow_html=True)

    if not st.session_state.edits_log:
        st.markdown("<div class='ibox'>Sin ediciones manuales en esta sesión. "
                    "La planificación refleja el resultado automático.</div>",
                    unsafe_allow_html=True)
    else:
        for i, log_e in enumerate(reversed(st.session_state.edits_log), 1):
            if log_e['tipo'] == 'individual':
                upms_str = ", ".join(log_e['upms'][:5])
                if len(log_e['upms']) > 5:
                    upms_str += f" (+{len(log_e['upms'])-5} más)"
                cambios_str = " · ".join(f"{k}→{v}" for k, v in log_e['cambios'].items())
                st.markdown(f"<div class='edit-card'>"
                            f"<b>#{len(st.session_state.edits_log)-i+1}</b> Individual · "
                            f"UPMs: {upms_str}<br>"
                            f"<span style='font-size:11px'>{cambios_str}</span></div>",
                            unsafe_allow_html=True)
            else:
                cambios_str = " · ".join(f"{k}→{v}" for k, v in log_e['cambios'].items())
                st.markdown(f"<div class='edit-card'>"
                            f"<b>#{len(st.session_state.edits_log)-i+1}</b> Lote · "
                            f"{log_e['n_upms']} UPMs desde {log_e['origen']}<br>"
                            f"<span style='font-size:11px'>{cambios_str}</span></div>",
                            unsafe_allow_html=True)

    # ── Resumen post-edición ──────────────────────────────────────────────────
    if st.session_state.edit_counter > 0:
        st.markdown("<div class='stitle'>Estado actual tras ediciones</div>",
                    unsafe_allow_html=True)
        df_post = df_plan[~df_plan['equipo'].isin(['sin_asignar'])].copy()
        res_post = df_post.groupby(['equipo', 'jornada']).agg(
            n_upms=('id_entidad', 'count'),
            viv=('viv', 'sum'),
            carga=('carga_pond', 'sum')
        ).reset_index()
        st.dataframe(res_post, use_container_width=True, hide_index=True)


# ══ TAB 4 — PLANIFICACIÓN DE JORNADAS ═════════
with tab_plan:
    st.markdown("<div class='stitle'>Calendario de Jornadas del Operativo</div>",
                unsafe_allow_html=True)
    st.markdown("""<div class='ibox'>
    Configura el estado de cada jornada. Si una jornada se <b>traslada</b>, el sistema
    detecta automáticamente si en el mes receptor habrá <b>3 jornadas</b> activas
    (la trasladada + las 2 propias) y lo señala en el calendario.<br>
    Las fechas de inicio que ingreses aquí se usarán en el Excel.
    </div>""",unsafe_allow_html=True)

    mes_ini_cal=st.session_state.mes_inicio_cal
    total_meses=int(df['mes'].max()) if df is not None and len(df)>0 else 12
    meses_todos=sorted(st.session_state.data_raw["mes"].dropna().unique().tolist())
    total_meses_op=len(meses_todos)

    cfg_j=st.session_state.config_jornadas

    all_jornadas=[]
    for mes in meses_todos:
        j1_n_,j2_n_,mes_n_=jornada_num_desde_mes(int(mes),mes_ini_cal)
        all_jornadas.append({'jn':j1_n_,'mes':mes,'mes_nombre':mes_n_,'mitad':1})
        all_jornadas.append({'jn':j2_n_,'mes':mes,'mes_nombre':mes_n_,'mitad':2})

    trasladadas_a={}
    for jinfo in all_jornadas:
        jn=jinfo['jn']
        estado=cfg_j.get(jn,{}).get('estado',ESTADO_OK)
        tr_a=cfg_j.get(jn,{}).get('trasladada_a',None)
        if estado==ESTADO_MV and tr_a:
            trasladadas_a.setdefault(tr_a,[]).append(jn)

    st.markdown(f"**{len(all_jornadas)} jornadas en {total_meses_op} meses operativos**")

    for mes in meses_todos:
        j1_n_,j2_n_,mes_n_=jornada_num_desde_mes(int(mes),mes_ini_cal)
        js_extra=trasladadas_a.get(int(mes),[])
        n_total_mes=2+len(js_extra)
        alerta_3=""
        if n_total_mes>=3:
            alerta_3=f" ⚠️ **{n_total_mes} jornadas** en este mes (incluye traslado)"

        with st.expander(f"📅 Mes {int(mes)} — {mes_n_} · J{j1_n_} + J{j2_n_}{alerta_3}",
                         expanded=(int(mes)==int(df['mes'].iloc[0]))):
            if js_extra:
                for jn_extra in js_extra:
                    st.markdown(f"<div class='jplan-mv'>🔀 Jornada {jn_extra} trasladada a este mes</div>",
                                unsafe_allow_html=True)
            for j_n_iter in [j1_n_,j2_n_]:
                st.markdown(f"**Jornada {j_n_iter}** (mitad {'1ª' if j_n_iter==j1_n_ else '2ª'})")
                cfg_this=cfg_j.get(j_n_iter,{})
                col_e,col_f,col_t=st.columns([2,2,2])
                with col_e:
                    estado_sel=st.selectbox(
                        "Estado",
                        [ESTADO_OK,ESTADO_MV,ESTADO_CAN],
                        index=[ESTADO_OK,ESTADO_MV,ESTADO_CAN].index(
                            cfg_this.get('estado',ESTADO_OK)),
                        key=f"est_{j_n_iter}")
                with col_f:
                    fecha_sel=st.date_input(
                        "Fecha de inicio",
                        value=cfg_this.get('fecha',None) or date.today(),
                        key=f"fec_{j_n_iter}")
                with col_t:
                    if estado_sel==ESTADO_MV:
                        meses_destino=[m for m in meses_todos if int(m)!=int(mes)]
                        tr_default=cfg_this.get('trasladada_a',
                                                 int(meses_destino[0]) if meses_destino else None)
                        tr_idx=0
                        if tr_default and tr_default in [int(m) for m in meses_destino]:
                            tr_idx=[int(m) for m in meses_destino].index(tr_default)
                        tr_mes=st.selectbox(
                            "Trasladar a mes",
                            [int(m) for m in meses_destino],
                            index=tr_idx,
                            format_func=lambda x: f"Mes {x} — {jornada_num_desde_mes(x,mes_ini_cal)[2]}",
                            key=f"tr_{j_n_iter}")
                    else:
                        tr_mes=None
                        st.markdown("<div style='height:38px'></div>",unsafe_allow_html=True)

                cfg_j[j_n_iter]={'estado':estado_sel,'fecha':fecha_sel,'trasladada_a':tr_mes}

                if estado_sel==ESTADO_OK:
                    st.markdown(f"<div class='jplan-ok'>✅ J{j_n_iter} — "
                                f"inicio {fecha_sel.strftime('%d/%m/%Y')} — "
                                f"{p['dias_op']} días → fin "
                                f"{(fecha_sel+timedelta(days=p['dias_op']-1)).strftime('%d/%m/%Y')}"
                                f"</div>",unsafe_allow_html=True)
                elif estado_sel==ESTADO_MV:
                    st.markdown(f"<div class='jplan-mv'>🔀 J{j_n_iter} trasladada → "
                                f"Mes {tr_mes} · se encuestará en ese período</div>",
                                unsafe_allow_html=True)
                else:
                    st.markdown(f"<div class='jplan-can'>❌ J{j_n_iter} cancelada</div>",
                                unsafe_allow_html=True)

    st.session_state.config_jornadas=cfg_j

    st.markdown("<div class='stitle'>Resumen del cronograma</div>",unsafe_allow_html=True)
    filas_res=[]
    for jinfo in all_jornadas:
        jn=jinfo['jn']; cfg_this=cfg_j.get(jn,{})
        estado=cfg_this.get('estado',ESTADO_OK)
        fecha=cfg_this.get('fecha',None)
        fecha_str=fecha.strftime("%d/%m/%Y") if fecha else "—"
        fin_str=(fecha+timedelta(days=p['dias_op']-1)).strftime("%d/%m/%Y") if fecha else "—"
        filas_res.append({
            'Jornada':f"J{jn}",'Mes':f"{int(jinfo['mes'])} — {jinfo['mes_nombre']}",
            'Estado':estado,'Inicio':fecha_str,'Fin':fin_str
        })
    st.dataframe(pd.DataFrame(filas_res),use_container_width=True,hide_index=True)

# ══ TAB 5 — REPORTE Y DESCARGA ════════════════
with tab_reporte:
    st.markdown("<div class='stitle'>Reporte y Descarga Excel</div>",unsafe_allow_html=True)
    st.markdown("""<div class='ibox'>
    El Excel incluye una hoja por <b>jornada planificada o trasladada</b>.
    Las jornadas canceladas se omiten. El número de jornada en cada hoja
    corresponde al número real del cronograma INEC.
    </div>""",unsafe_allow_html=True)

    # Indicar si hay ediciones manuales
    if st.session_state.edit_counter > 0:
        st.markdown(f"<div class='edit-ok'>✏️ El Excel incluirá "
                    f"{st.session_state.edit_counter} edición(es) manual(es).</div>",
                    unsafe_allow_html=True)

    # Catálogo territorial
    st.markdown("<div class='stitle'>Catálogo territorial para completar el Excel</div>",
                unsafe_allow_html=True)
    st.markdown("""<div class='ibox'>
    Sube el Excel o CSV con la organización territorial del Ecuador. La app usa
    ese catálogo para completar provincia, cantón, parroquia y códigos auxiliares
    en el reporte final.
    </div>""", unsafe_allow_html=True)

    cat_file = st.file_uploader(
        "Catálogo territorial (.xlsx, .xls o .csv)",
        type=["xlsx", "xls", "csv"],
        key="catalogo_territorial_up"
    )
    if cat_file is not None:
        try:
            df_cat = cargar_catalogo_territorial(cat_file)
            lookup_cat, cols_cat = preparar_lookup_territorial(df_cat)
            st.session_state.catalogo_df = df_cat
            st.session_state.catalogo_lookup = lookup_cat
            st.session_state.catalogo_cols = cols_cat
            st.success(f"✓ Catálogo cargado: {len(df_cat):,} filas")
        except Exception as e:
            st.error(f"No se pudo leer el catálogo territorial: {e}")

    if st.session_state.catalogo_df is not None:
        cols_detectadas = {k: v for k, v in st.session_state.catalogo_cols.items() if v}
        st.caption(f"Columnas detectadas: {cols_detectadas}")

    # Personal por equipo
    st.markdown("<div class='stitle'>Personal por equipo</div>",unsafe_allow_html=True)
    for eq in eq_cfg:
        nombre_eq=eq["nombre"]; n_enc_eq=eq["enc"]
        pi_prev=st.session_state.personal_info.get(nombre_eq,{})
        with st.expander(f"👥 {nombre_eq} — {n_enc_eq} enc.",expanded=False):
            pc1,pc2,pc3=st.columns(3)
            with pc1: sup_n=st.text_input("Supervisor",value=pi_prev.get('supervisor_nombre',''),key=f"sup_n_{nombre_eq}")
            with pc2: sup_c=st.text_input("Cédula sup.",value=pi_prev.get('supervisor_cedula',''),key=f"sup_c_{nombre_eq}")
            with pc3: sup_t=st.text_input("Celular sup.",value=pi_prev.get('supervisor_celular',''),key=f"sup_t_{nombre_eq}")
            enc_list_new=[]
            for j in range(n_enc_eq):
                prev_enc=pi_prev.get('encuestadores',[{}]*n_enc_eq)
                prev_j=prev_enc[j] if j<len(prev_enc) else {}
                pe1,pe2,pe3=st.columns(3)
                with pe1: en=st.text_input(f"Encuestador {j+1}",value=prev_j.get('nombre',''),key=f"enc_n_{nombre_eq}_{j}")
                with pe2: ec=st.text_input(f"Cédula {j+1}",value=prev_j.get('cedula',''),key=f"enc_c_{nombre_eq}_{j}")
                with pe3: et=st.text_input(f"Celular {j+1}",value=prev_j.get('celular',''),key=f"enc_t_{nombre_eq}_{j}")
                enc_list_new.append({'nombre':en,'cedula':ec,'celular':et,'cod':''})
            pch1,pch2=st.columns(2)
            with pch1: ch_n=st.text_input("Chofer",value=pi_prev.get('chofer_nombre',''),key=f"ch_n_{nombre_eq}")
            with pch2: plca=st.text_input("Placa",value=pi_prev.get('placa',''),key=f"plca_{nombre_eq}")
            st.session_state.personal_info[nombre_eq]={
                'supervisor_nombre':sup_n,'supervisor_cedula':sup_c,
                'supervisor_celular':sup_t,'supervisor_cod':'',
                'encuestadores':enc_list_new,'n_enc':n_enc_eq,
                'chofer_nombre':ch_n,'placa':plca,
            }

    # Resumen
    st.markdown("<div class='stitle'>Resumen</div>",unsafe_allow_html=True)
    if res_bal is not None and len(res_bal)>0:
        tr=pd.DataFrame([{'equipo':'TOTAL','jornada':'—',
            'n_upms':res_bal['n_upms'].sum(),
            'viv_reales':res_bal['viv_reales'].sum(),
            'carga_ponderada':res_bal['carga_ponderada'].sum(),
            'dist_km':res_bal.get('dist_km',pd.Series([0])).sum()}])
        st.dataframe(pd.concat([res_bal,tr],ignore_index=True).rename(columns={
            'equipo':'Equipo','jornada':'Jornada','n_upms':'UPMs',
            'viv_reales':'Viv.','carga_ponderada':'Carga pond.','dist_km':'Dist (km)'}),
            use_container_width=True)

    cols_ok=[c for c in ['id_entidad','tipo_entidad','viv','carga_pond',
                          'equipo','jornada','encuestador','dia_inicio','dia_fin']
             if c in df_plan.columns]
    st.dataframe(df_plan[cols_ok].sort_values(
        ['equipo','jornada','encuestador','dia_inicio']).reset_index(drop=True),
        use_container_width=True,height=280)

    # Descarga Excel
    st.markdown("<div class='stitle'>Descargar Excel</div>",unsafe_allow_html=True)

    cfg_j=st.session_state.config_jornadas
    mes_ini_cal=st.session_state.mes_inicio_cal

    jornadas_excel=[]
    for jinfo in all_jornadas if 'all_jornadas' in dir() else []:
        jn=jinfo['jn']; cfg_this=cfg_j.get(jn,{})
        estado=cfg_this.get('estado',ESTADO_OK)
        if estado==ESTADO_CAN: continue
        jornadas_excel.append({
            'jornada_num':jn,
            'jornada_nombre':'Jornada 1' if jinfo['mitad']==1 else 'Jornada 2',
            'fecha':cfg_this.get('fecha',None)
        })
    if not jornadas_excel:
        j1_n_,j2_n_,_=jornada_num_desde_mes(int(df['mes'].iloc[0]),mes_ini_cal)
        cfg_j_=st.session_state.config_jornadas
        for jn_,jnom_ in [(j1_n_,'Jornada 1'),(j2_n_,'Jornada 2')]:
            jornadas_excel.append({
                'jornada_num':jn_,'jornada_nombre':jnom_,
                'fecha':cfg_j_.get(jn_,{}).get('fecha',None)
            })

    mes_n_excel=jornada_num_desde_mes(int(df['mes'].iloc[0]),mes_ini_cal)[2]

    if st.button("📋 Generar Excel",use_container_width=True,type="primary"):
        with st.spinner("Generando Excel..."):
            try:
                excel_bytes=generar_excel(
                    df_plan=df_plan,eq_cfg=eq_cfg,
                    personal_info=st.session_state.personal_info,
                    jornadas_activas=jornadas_excel,
                    dias_op=p["dias_op"],
                    catalogo_lookup = st.session_state.get('catalogo_lookup', {}))

                nums=[str(ji['jornada_num']) for ji in jornadas_excel]
                fname=f"planificacion_J{'_'.join(nums)}_{mes_n_excel}.xlsx"
                st.download_button(
                    label=f"⬇️ Descargar {fname}",data=excel_bytes,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
                st.success("✓ Excel listo.")
            except Exception as e:
                st.error(f"Error: {e}")
                import traceback; st.code(traceback.format_exc())
