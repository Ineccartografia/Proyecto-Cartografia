# App fusionada generada a partir de app-1.py y app_modificado-1.py.
# Base visual: app_modificado
# Ajustes funcionales solicitados:
# - Mantener edición manual
# - Mantener estética de app_modificado
# - Catálogo territorial desde .txt público de GitHub
# - Quitar pestaña de planificación de jornadas
# - Reporte basado en mes operativo seleccionado

import streamlit as st
import pandas as pd
import numpy as np
import geopandas as gpd
import folium
import pyogrio
from streamlit_folium import st_folium
import plotly.express as px
import tempfile, os, warnings, io, requests
from io import StringIO
from datetime import date, timedelta
import osmnx as ox
import networkx as nx
from networkx.algorithms import approximation
from pyproj import Transformer
from sklearn.cluster import KMeans
from sklearn.metrics import silhouette_score
from sklearn.neighbors import BallTree
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')

st.set_page_config(page_title="INEC · ENDI Planificación", page_icon="📊", layout="wide", initial_sidebar_state="expanded")

INEC_LOGO = "https://upload.wikimedia.org/wikipedia/commons/a/a8/Logo_del_INEC_Ecuador.png"
BASE_LAT = -2.145825935522539
BASE_LON = -79.89383956329586
PRO_GYE = "09"
CAN_GYE = "01"
MESES_CAL = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}
COLORES = ['#dc2626','#003B71','#059669','#d97706','#7c3aed','#0891b2','#c2410c','#be185d']

st.markdown(f"""
<style>
.block-container {{padding-top: 1.2rem; padding-bottom: 2rem; max-width: 96%;}}
.stApp {{background: linear-gradient(180deg, #f8fafc 0%, #eef4f8 100%);}}
.hdr {{display:flex; gap:16px; align-items:center; background:#ffffff; border:1px solid #d7e3ee; padding:18px 22px; border-radius:18px; box-shadow:0 8px 24px rgba(0,59,113,.08); margin-bottom:16px;}}
.hdr img {{width:58px; height:auto;}}
.hdr-text h1 {{margin:0; color:#003B71; font-size:1.6rem;}}
.hdr-text p {{margin:.2rem 0 0 0; color:#53657a;}}
.stitle {{font-weight:700; color:#003B71; margin:.6rem 0 .45rem 0; font-size:1.05rem;}}
.ibox {{background:#ffffff; border:1px solid #d8e2ec; border-left:4px solid #003B71; padding:12px 14px; border-radius:12px; color:#34495e; margin:.5rem 0;}}
.wbox {{background:#fff7ed; border:1px solid #fed7aa; color:#9a3412; padding:10px 12px; border-radius:10px;}}
.kcard {{background:#ffffff; border:1px solid #dbe6ef; border-radius:14px; padding:14px; text-align:center; box-shadow:0 4px 18px rgba(0,0,0,.04);}}
.kcard .v {{font-size:1.35rem; font-weight:700;}}
.kcard .l {{font-size:.86rem; color:#516579; font-weight:600;}}
.kcard .s {{font-size:.72rem; color:#7b8b9a;}}
.edit-ok {{background:#ecfdf5; color:#065f46; border:1px solid #a7f3d0; padding:10px 12px; border-radius:10px; margin:.4rem 0;}}
.edit-card {{background:#fff; border:1px solid #dbe6ef; border-radius:12px; padding:10px 12px; margin:.35rem 0;}}
.balance-box, .eq-card, .bcard {{background:#fff; border:1px solid #dbe6ef; border-radius:14px; padding:12px;}}
.sidebar-logo {{display:flex; gap:12px; align-items:center;}}
.sidebar-logo img {{width:42px;}}
.sidebar-title {{font-weight:700; color:#003B71;}}
.sidebar-sub {{font-size:.84rem; color:#64748b;}}
</style>
""", unsafe_allow_html=True)

def cv_pct(s):
    m = s.mean()
    return float(s.std()/m*100) if m > 0 else 0.0

def utm_to_wgs84(df):
    t = Transformer.from_crs("epsg:32717","epsg:4326",always_xy=True)
    lons,lats = t.transform(df["x"].values,df["y"].values)
    df = df.copy(); df["lon"]=lons; df["lat"]=lats
    return df

def parse_codigo(codigo):
    c = str(codigo).strip()
    r = {'prov':'','canton':'','ciudad_parroq':'','zona':'','sector':'','man':''}
    if len(c)>=6: r['prov']=c[:2]; r['canton']=c[2:4]; r['ciudad_parroq']=c[4:6]
    if len(c)>=9: r['zona']=c[6:9]
    if len(c)>=12: r['sector']=c[9:12]
    if len(c)>=15: r['man']=c[12:15]
    return r

def normalizar_codigo(valor, ancho=None):
    if pd.isna(valor): return ''
    s = str(valor).strip()
    if s.endswith('.0'): s = s[:-2]
    s = ''.join(ch for ch in s if ch.isdigit())
    if not s: return ''
    if ancho: s = s.zfill(ancho)[-ancho:]
    return s

def detectar_columnas_catalogo(df_cat):
    cols = {str(c).strip().upper(): c for c in df_cat.columns}
    def pick(*cands):
        for cand in cands:
            if cand in cols: return cols[cand]
        return None
    return {
        'parroquia_cod': pick('DPA_PARROQ', 'COD_PARROQUIA', 'PARROQUIA_CODIGO'),
        'parroquia_nom': pick('DPA_DESPAR', 'DESC_PARROQUIA', 'PARROQUIA'),
        'canton_cod' : pick('DPA_CANTON', 'COD_CANTON', 'CANTON_CODIGO'),
        'canton_nom' : pick('DPA_DESCAN', 'DESC_CANTON', 'CANTON'),
        'prov_cod' : pick('DPA_PROVIN', 'COD_PROVINCIA', 'PROVINCIA_CODIGO'),
        'prov_nom' : pick('DPA_DESPRO', 'DESC_PROVINCIA', 'PROVINCIA'),
        'tipo_txt' : pick('TXT', 'TIPO', 'TIPO_SECTOR', 'CLASE'),
        'fcode' : pick('FCODE', 'COD_FCODE', 'CODIGO_FCODE')
    }

def cargar_catalogo_github_txt(url):
    r = requests.get(url, timeout=30)
    r.raise_for_status()
    txt = r.text
    candidatos = ['|', '\t', ';', ',']
    mejor_df, mejor_cols = None, 0
    for sep in candidatos:
        try:
            df = pd.read_csv(StringIO(txt), sep=sep, dtype=str)
            if len(df.columns) > mejor_cols:
                mejor_df, mejor_cols = df.copy(), len(df.columns)
        except Exception:
            pass
    if mejor_df is None:
        raise ValueError('No se pudo interpretar el .txt del catálogo territorial')
    mejor_df.columns = [str(c).strip() for c in mejor_df.columns]
    return mejor_df

def preparar_lookup_territorial(df_cat):
    if df_cat is None or len(df_cat) == 0: return {}, {}
    cols = detectar_columnas_catalogo(df_cat)
    cod_col = cols.get('parroquia_cod')
    if not cod_col: return {}, cols
    work = df_cat.copy()
    work['__parroq_cod__'] = work[cod_col].apply(lambda v: normalizar_codigo(v, 6))
    work = work[work['__parroq_cod__'] != ''].drop_duplicates('__parroq_cod__', keep='first')
    lookup = {}
    for _, row in work.iterrows():
        cod = row['__parroq_cod__']
        lookup[cod] = {
            'provincia_codigo': normalizar_codigo(row[cols['prov_cod']], 2) if cols.get('prov_cod') else cod[:2],
            'provincia_nombre': str(row[cols['prov_nom']]).strip() if cols.get('prov_nom') and pd.notna(row[cols['prov_nom']]) else '',
            'canton_codigo': normalizar_codigo(row[cols['canton_cod']], 4) if cols.get('canton_cod') else cod[:4],
            'canton_nombre': str(row[cols['canton_nom']]).strip() if cols.get('canton_nom') and pd.notna(row[cols['canton_nom']]) else '',
            'parroquia_codigo': cod,
            'parroquia_nombre': str(row[cols['parroquia_nom']]).strip() if cols.get('parroquia_nom') and pd.notna(row[cols['parroquia_nom']]) else '',
            'tipo_txt': str(row[cols['tipo_txt']]).strip() if cols.get('tipo_txt') and pd.notna(row[cols['tipo_txt']]) else '',
            'fcode': str(row[cols['fcode']]).strip() if cols.get('fcode') and pd.notna(row[cols['fcode']]) else ''
        }
    return lookup, cols

def enriquecer_plan_con_catalogo(df_plan, catalogo_lookup):
    if df_plan is None or len(df_plan) == 0: return df_plan
    if not catalogo_lookup: return df_plan.copy()
    df_out = df_plan.copy()
    provs, cants, parroqs, tipos = [], [], [], []
    for _, row in df_out.iterrows():
        partes = parse_codigo(row.get('id_entidad', ''))
        cod_parr = f"{partes['prov']}{partes['canton']}{partes['ciudad_parroq']}"
        geo = catalogo_lookup.get(cod_parr, {})
        provs.append(geo.get('provincia_nombre', ''))
        cants.append(geo.get('canton_nombre', ''))
        parroqs.append(geo.get('parroquia_nombre', ''))
        tipos.append(geo.get('tipo_txt', ''))
    df_out['provincia_nombre'] = provs
    df_out['canton_nombre'] = cants
    df_out['parroquia_nombre'] = parroqs
    df_out['tipo_asentamiento'] = tipos
    return df_out

def jornada_num_desde_mes(mes_operativo, mes_inicio_cal):
    mes_cal = ((mes_inicio_cal - 1 + mes_operativo - 1) % 12) + 1
    j1_num = (mes_operativo - 1) * 2 + 1
    j2_num = j1_num + 1
    return j1_num, j2_num, MESES_CAL[mes_cal]

def cargar_gpkg(path, dissolve_upm=True):
    capas = pyogrio.list_layers(path)
    gdf = gpd.read_file(path, layer=capas[0][0])
    col_map = {'1_mes_cart':'mes','viv_total':'viv','1_zonal':'zonal','1_id_upm':'upm','ManSec':'id_entidad'}
    gdf = gdf.rename(columns={k:v for k,v in col_map.items() if k in gdf.columns})
    if 'id_entidad' in gdf.columns:
        gdf['tipo_entidad'] = gdf['id_entidad'].astype(str).apply(lambda x: 'sec' if '999' in x else 'man')
    gdf_u = gdf.to_crs(epsg=32717)
    gdf_u['geometry'] = gdf_u.geometry.representative_point()
    gdf_u['x'] = gdf_u.geometry.x; gdf_u['y'] = gdf_u.geometry.y
    if 'pro' in gdf_u.columns: gdf_u['pro_x'] = gdf_u['pro']
    if 'can' in gdf_u.columns: gdf_u['can_x'] = gdf_u['can']
    if 'mes' in gdf_u.columns: gdf_u['mes'] = pd.to_numeric(gdf_u['mes'],errors='coerce')
    return utm_to_wgs84(gdf_u)

def clustering_balanceado(df, n_clusters, cv_objetivo=0.20, max_iter=80, k_vecinos=15):
    coords = df[['x','y']].values.astype(float)
    cargas = df['carga_pond'].values.astype(float)
    km = KMeans(n_clusters=n_clusters, init='k-means++', n_init=10, max_iter=300, random_state=42)
    labels = km.fit_predict(coords).copy()
    def cluster_sums(): return np.array([cargas[labels==c].sum() for c in range(n_clusters)])
    cv_ini = cv_pct(pd.Series(cluster_sums()))
    log = [{'iter':0,'cv':cv_ini,'modo':'inicial'}]
    sin_mejora = 0
    for it in range(1, max_iter+1):
        sums = cluster_sums(); media = sums.mean(); cv = cv_pct(pd.Series(sums))
        if media > 0 and (np.abs(sums - media) / media).max() <= cv_objetivo:
            log.append({'iter':it,'cv':cv,'modo':'objetivo alcanzado ✓'})
            break
        k_max = int(sums.argmax()); k_min = int(sums.argmin())
        mask_max = np.where(labels == k_max)[0]; mask_min = np.where(labels == k_min)[0]
        if len(mask_max) == 0 or len(mask_min) == 0: break
        cent_min = coords[mask_min].mean(axis=0)
        dists = np.linalg.norm(coords[mask_max] - cent_min, axis=1)
        orden = np.argsort(dists)
        mejora = False
        for pos in orden[:15]:
            idx = mask_max[pos]
            labels[idx] = k_min
            new_cv = cv_pct(pd.Series(cluster_sums()))
            if new_cv < cv:
                log.append({'iter':it,'cv':new_cv,'modo':'rebalanceo'})
                mejora = True; sin_mejora = 0; break
            labels[idx] = k_max
        if not mejora:
            sin_mejora += 1
            log.append({'iter':it,'cv':cv,'modo':'sin mejora'})
            if sin_mejora >= 10: break
    cv_fin = cv_pct(pd.Series(cluster_sums()))
    log.append({'iter':len(log),'cv':cv_fin,'modo':'final'})
    return labels, log, cv_ini, cv_fin

def nearest_neighbor_order(points_xy, start_xy=None):
    n = len(points_xy)
    if n == 0: return []
    if n == 1: return [0]
    visited = [False]*n
    cur = int(np.argmin(np.linalg.norm(points_xy-start_xy,axis=1))) if start_xy is not None else 0
    order=[cur]; visited[cur]=True
    for _ in range(n-1):
        best_d,best_j=np.inf,-1; px,py=points_xy[cur]
        for j in range(n):
            if not visited[j]:
                d=(points_xy[j,0]-px)**2+(points_xy[j,1]-py)**2
                if d < best_d: best_d,best_j=d,j
        visited[best_j]=True; order.append(best_j); cur=best_j
    return order

def asignar_encuestadores_y_dias(df_grp, n_enc, dias_tot, viv_min, viv_max, inicio_dia=1):
    target = (viv_min + viv_max) / 2.0
    ultimo = inicio_dia + dias_tot - 1
    df_g = df_grp.copy()
    n_rows = len(df_g)
    cargas = df_g['carga_pond'].values.astype(float)
    viviendas = df_g['viv'].values.astype(float)
    coords = df_g[['x','y']].values.astype(float)
    orden_bp = np.argsort(cargas)[::-1]
    enc_acum = np.zeros(n_enc)
    enc_asig = np.zeros(n_rows, dtype=int)
    for pos in orden_bp:
        em = int(np.argmin(enc_acum)); enc_asig[pos] = em + 1; enc_acum[em] += cargas[pos]
    geo_order = []
    for enc_id in range(1, n_enc+1):
        pos_enc = np.where(enc_asig == enc_id)[0]
        if len(pos_enc)==0: continue
        sub_coords = coords[pos_enc]; centroide = sub_coords.mean(axis=0)
        nn = nearest_neighbor_order(sub_coords, start_xy=centroide)
        for nn_pos in nn: geo_order.append((pos_enc[nn_pos], enc_id))
    dias_range = list(range(inicio_dia, ultimo+1))
    calendario = {e:{d:0.0 for d in dias_range} for e in range(1,n_enc+1)}
    cursor = {e:inicio_dia for e in range(1,n_enc+1)}
    dia_ini_arr = np.full(n_rows, inicio_dia, dtype=int)
    dia_fin_arr = np.full(n_rows, inicio_dia, dtype=int)
    for pos, enc_id in geo_order:
        viv_m = max(0.0, viviendas[pos]); cal = calendario[enc_id]; cur = cursor[enc_id]
        if viv_m > viv_max:
            dias_m = max(1, int(np.ceil(viv_m / target))); dias_m = min(dias_m, dias_tot); bloque = None
            for d_s in range(cur, ultimo - dias_m + 2):
                if all(cal.get(d, target) < target for d in range(d_s, d_s+dias_m)):
                    bloque = d_s; break
            if bloque is None: bloque = max(inicio_dia, min(cur, ultimo - dias_m + 1))
            d_ini = bloque; d_fin = min(d_ini + dias_m - 1, ultimo); vpd = viv_m / max(1, d_fin - d_ini + 1)
            for dd in range(d_ini, d_fin+1): cal[dd] = cal.get(dd, 0.0) + vpd
            cursor[enc_id] = d_fin + 1
        else:
            dia_asig = None
            for d in range(cur, ultimo+1):
                if cal.get(d, 0.0) < target: dia_asig = d; break
            if dia_asig is None: dia_asig = ultimo
            cal[dia_asig] = cal.get(dia_asig, 0.0) + viv_m
            d_ini = dia_asig; d_fin = dia_asig
            if cal[dia_asig] >= target and cursor[enc_id] == dia_asig: cursor[enc_id] = min(dia_asig+1, ultimo)
        dia_ini_arr[pos] = max(inicio_dia, min(d_ini, ultimo))
        dia_fin_arr[pos] = max(dia_ini_arr[pos], min(d_fin, ultimo))
    df_g['encuestador'] = enc_asig
    df_g['dia_inicio'] = dia_ini_arr
    df_g['dia_fin'] = dia_fin_arr
    df_g['dia_operativo'] = dia_ini_arr
    return df_g

def generar_excel(df_plan, eq_cfg, personal_info, jornadas_activas, dias_op, catalogo_lookup=None):
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    df_plan_exp = enriquecer_plan_con_catalogo(df_plan, catalogo_lookup)
    AZ_OSCURO="0D3B6E"; AZ_MEDIO="1A5276"; AZ_CLARO="D6EAF8"; VRD_CHECK="D5F5E3"; BLANCO="FFFFFF"
    ENC_PALETAS = [{"par":"DBEAFE","impar":"EFF6FF","subtot":"BFDBFE","hdr":"1D4ED8"},{"par":"D1FAE5","impar":"ECFDF5","subtot":"A7F3D0","hdr":"065F46"},{"par":"FEF9C3","impar":"FEFCE8","subtot":"FDE68A","hdr":"854D0E"},{"par":"FCE7F3","impar":"FDF4FF","subtot":"F9A8D4","hdr":"831843"}]
    def sc(cell,bold=False,bg=None,fg="000000",ha="left",sz=9,brd=False,wrap=False):
        cell.font=Font(bold=bold,size=sz,color=fg); cell.alignment=Alignment(horizontal=ha,vertical="center",wrap_text=wrap)
        if bg: cell.fill=PatternFill("solid",fgColor=bg)
        if brd:
            t=Side(style='thin'); cell.border=Border(left=t,right=t,top=t,bottom=t)
    for jinfo in jornadas_activas:
        j_num = jinfo['jornada_num']; j_nombre_hoja = jinfo['jornada_nombre']; fecha_inicio = jinfo.get('fecha')
        df_jor = df_plan_exp[df_plan_exp['jornada']==j_nombre_hoja].copy()
        if len(df_jor)==0: continue
        ws = wb.create_sheet(title=f"J{j_num}")
        ws.sheet_view.showGridLines=False
        cur=1; last_col=15+dias_op
        ws.merge_cells(f'A{cur}:{get_column_letter(last_col)}{cur}'); ws.cell(cur,1,'INSTITUTO NACIONAL DE ESTADÍSTICA Y CENSOS'); sc(ws.cell(cur,1),bold=True,bg=AZ_OSCURO,fg=BLANCO,ha='center'); cur+=1
        ws.merge_cells(f'A{cur}:{get_column_letter(last_col)}{cur}'); ws.cell(cur,1,'PROGRAMACIÓN OPERATIVO DE CAMPO'); sc(ws.cell(cur,1),bold=True,bg=AZ_OSCURO,fg=BLANCO,ha='center'); cur+=2
        if fecha_inicio:
            fechas=[fecha_inicio+timedelta(days=i) for i in range(dias_op)]
        else:
            fechas=None
        equipos_jor=[e['nombre'] for e in eq_cfg if e['nombre'] in df_jor['equipo'].values]
        for grupo_num,nombre_eq in enumerate(equipos_jor,1):
            dfeq=df_jor[df_jor['equipo']==nombre_eq].copy()
            if len(dfeq)==0: continue
            sc(ws.cell(cur,1,'JORNADA'),bold=True); ws.cell(cur,2,str(j_num)); sc(ws.cell(cur,7,'GRUPO'),bold=True); ws.cell(cur,8,str(grupo_num)); cur+=2
            hdrs=['SUPERVISOR','ENCUESTADOR','CARGA','PROV','CANTON','CIUDAD/PARROQ','ZONA','SECTOR','MAN','CÓDIGO','PROVINCIA','CANTÓN','CIUDAD','NRO EDIF']
            for ci,h in enumerate(hdrs,1): sc(ws.cell(cur,ci,h),bold=True,bg=AZ_CLARO,ha='center',sz=8,brd=True,wrap=True)
            for i in range(dias_op): sc(ws.cell(cur,15+i, fechas[i].strftime('%d/%m') if fechas else f'D{i+1}'),bold=True,bg=AZ_CLARO,ha='center',sz=8,brd=True)
            sc(ws.cell(cur,15+dias_op,'VIV'),bold=True,bg=AZ_CLARO,ha='center',sz=8,brd=True); cur+=1
            dfeq = dfeq.sort_values(['encuestador','dia_inicio'])
            for ridx,(_,rd) in enumerate(dfeq.iterrows()):
                pal = ENC_PALETAS[int(rd.get('encuestador',1)-1)%len(ENC_PALETAS)]
                bg = pal['par'] if ridx%2==0 else pal['impar']
                partes = parse_codigo(str(rd.get('id_entidad','')))
                vals=[ '', int(rd.get('encuestador',0)), round(float(rd.get('carga_pond',0)),1), partes['prov'], partes['canton'], partes['ciudad_parroq'], partes['zona'], partes['sector'], partes['man'], str(rd.get('id_entidad','')), str(rd.get('provincia_nombre','')), str(rd.get('canton_nombre','')), str(rd.get('parroquia_nombre','')), str(rd.get('fcode','')) ]
                for ci,val in enumerate(vals,1): sc(ws.cell(cur,ci,val),bg=bg,ha='center',sz=8,brd=True)
                dini = int(rd.get('dia_inicio', rd.get('dia_operativo',1))); dfin = int(rd.get('dia_fin', dini))
                for i in range(dias_op): sc(ws.cell(cur,15+i,'') ,bg=(VRD_CHECK if dini <= i+1 <= dfin else bg),ha='center',brd=True)
                sc(ws.cell(cur,15+dias_op,int(rd.get('viv',0))),bg=bg,ha='center',sz=8,brd=True)
                cur += 1
            cur += 2
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

if 'data_raw' not in st.session_state:
    st.session_state.data_raw=None
    st.session_state.data_mes=None
    st.session_state.graph_G=None
    st.session_state.resultados_generados=False
    st.session_state.df_plan=None
    st.session_state.tsp_results={}
    st.session_state.road_paths={}
    st.session_state.resumen_bal=None
    st.session_state.sil_score=None
    st.session_state.nbombero=0
    st.session_state.personal_info={}
    st.session_state.balance_log=[]
    st.session_state.cv_ini_bal=None
    st.session_state.cv_fin_bal=None
    st.session_state.viv_por_cluster_antes=None
    st.session_state.viv_por_cluster_despues=None
    st.session_state.mes_inicio_cal=7
    st.session_state.params={"dias_op":12,"viv_min":50,"viv_max":80,"factor_r":1.5,"usar_bomb":True,"usar_gye":True,"dias_gye":3,"umbral_gye":10,"cv_objetivo":20,"max_iter_bal":80,"k_vecinos":15,"min_dist_bomb_m":40000}
    st.session_state.equipos_cfg=[{"id":1,"nombre":"Equipo 1","enc":3},{"id":2,"nombre":"Equipo 2","enc":3},{"id":3,"nombre":"Equipo 3","enc":3}]
    st.session_state.catalogo_df=None
    st.session_state.catalogo_lookup={}
    st.session_state.catalogo_cols={}
    st.session_state.editslog=[]
    st.session_state.editcounter=0
    st.session_state.github_txt_url=""

with st.sidebar:
    st.markdown(f'<div class="sidebar-logo"><img src="{INEC_LOGO}"><div><div class="sidebar-title">Encuesta Nacional</div><div class="sidebar-sub">INEC · Zonal Litoral</div></div></div>', unsafe_allow_html=True)
    st.divider()
    st.markdown('**PASO 1**')
    gpkg_f=st.file_uploader('GeoPackage',type=['gpkg'],key='gpkg_up')
    if gpkg_f and st.button('Procesar',use_container_width=True,type='primary'):
        with st.spinner('Leyendo geometrías...'):
            with tempfile.NamedTemporaryFile(delete=False,suffix='.gpkg') as tmp:
                tmp.write(gpkg_f.read()); p_tmp=tmp.name
            data=cargar_gpkg(p_tmp, dissolve_upm=True); os.unlink(p_tmp)
            st.session_state.data_raw=data; st.session_state.resultados_generados=False; st.success(f'{len(data):,} entidades')
    st.divider()
    st.markdown('**PASO 2**')
    gml_f=st.file_uploader('GraphML',type=['graphml'],key='gml_up')
    if gml_f and st.button('Cargar grafo',use_container_width=True):
        with tempfile.NamedTemporaryFile(delete=False,suffix='.graphml') as tmp:
            tmp.write(gml_f.read()); pg=tmp.name
        G=ox.load_graphml(pg); os.unlink(pg)
        st.session_state.graph_G=G; st.success(f'{len(G.nodes):,} nodos')
    st.divider()
    if st.session_state.data_raw is not None:
        meses_disp=sorted(st.session_state.data_raw['mes'].dropna().unique().tolist())
        mes_sel=st.selectbox('Mes operativo', meses_disp, format_func=lambda x: f'Mes {int(x)}')
        df_mes=st.session_state.data_raw[st.session_state.data_raw['mes']==mes_sel].copy(); st.session_state.data_mes=df_mes
        mes_ini_cal=st.selectbox('El mes operativo 1 corresponde a:', list(MESES_CAL.keys()), index=st.session_state.mes_inicio_cal-1, format_func=lambda x: MESES_CAL[x])
        st.session_state.mes_inicio_cal=mes_ini_cal
        j1_n,j2_n,mes_nom=jornada_num_desde_mes(int(mes_sel), mes_ini_cal)
        st.markdown(f'<div class="ibox">Mes {int(mes_sel)} ({mes_nom}) → <b>Jornada {j1_n}</b> + <b>Jornada {j2_n}</b></div>', unsafe_allow_html=True)
        st.divider()
        c1,c2=st.columns(2)
        with c1:
            if st.button('＋',use_container_width=True):
                nid=max(t['id'] for t in st.session_state.equipos_cfg)+1
                st.session_state.equipos_cfg.append({'id':nid,'nombre':f'Equipo {nid}','enc':3})
        with c2:
            if st.button('－',use_container_width=True, disabled=len(st.session_state.equipos_cfg)<=1):
                st.session_state.equipos_cfg.pop()
        for i,eq in enumerate(st.session_state.equipos_cfg):
            cc1,cc2=st.columns([2,1])
            with cc1:
                st.session_state.equipos_cfg[i]['nombre']=st.text_input(f"n{eq['id']}", value=eq['nombre'], key=f"n_{eq['id']}", label_visibility='collapsed')
            with cc2:
                st.session_state.equipos_cfg[i]['enc']=st.number_input('e', min_value=1, max_value=6, value=eq['enc'], key=f"e_{eq['id']}", label_visibility='collapsed')
        p=st.session_state.params
        p['dias_op']=st.slider('Días operativos',10,14,p['dias_op'])
        p['viv_min']=st.slider('Mín viv/día',30,60,p['viv_min'])
        p['viv_max']=st.slider('Máx viv/día',60,120,p['viv_max'])
        p['factor_r']=st.slider('Factor rural',1.0,2.5,p['factor_r'],0.1)
        p['cv_objetivo']=st.slider('Tolerancia de carga (%)',10,50,p['cv_objetivo'])
        p['max_iter_bal']=st.slider('Iter. máx.',20,200,p['max_iter_bal'],step=10)
        p['k_vecinos']=st.slider('Vecinos frontera (k)',5,30,p['k_vecinos'])

st.markdown(f'<div class="hdr"><img src="{INEC_LOGO}"><div class="hdr-text"><h1>Planificación Automática · Actualización Cartográfica</h1><p>Instituto Nacional de Estadística y Censos · Zonal Litoral · ENDI 2025</p></div></div>', unsafe_allow_html=True)

if st.session_state.data_raw is None:
    st.markdown('<div class="ibox">Carga el <code>.gpkg</code> desde el panel lateral.</div>', unsafe_allow_html=True)
    st.stop()

df=st.session_state.data_mes
if df is None or len(df)==0:
    st.warning('Sin datos para el mes seleccionado.'); st.stop()

p=st.session_state.params
k1,k2,k3,k4,k5=st.columns(5)
cv_v=cv_pct(df['viv']); cv_c='#059669' if cv_v<50 else '#dc2626'
for col,(val,lbl,sub,c) in zip([k1,k2,k3,k4,k5],[(f"{len(df):,}",'UPMs',f"mes {int(df['mes'].iloc[0])}",'#003B71'),(f"{int(df['viv'].sum()):,}",'Viviendas','precenso 2020','#003B71'),(f"{len(df[df['tipo_entidad'].astype(str).str.contains('man')]):,}",'Amanzanadas','man/man_upm','#003B71'),(f"{len(df[df['tipo_entidad'].astype(str).str.contains('sec')]):,}",'Dispersas','sec/sec_upm','#003B71'),(f"{cv_v:.1f}%",'CV viviendas','dispersión',cv_c)]):
    with col:
        st.markdown(f'<div class="kcard"><div class="v" style="color:{c}">{val}</div><div class="l">{lbl}</div><div class="s">{sub}</div></div>', unsafe_allow_html=True)

cb1,cb2=st.columns([1,3])
with cb1:
    btn=st.button('Generar Planificación', use_container_width=True, type='primary', disabled=(st.session_state.graph_G is None))
with cb2:
    if st.session_state.graph_G is None:
        st.markdown('<div class="wbox">Carga el <code>.graphml</code> en el Paso 2.</div>', unsafe_allow_html=True)
    elif st.session_state.resultados_generados:
        st.markdown('<div class="ibox">Planificación lista. Puedes regenerar o editar manualmente.</div>', unsafe_allow_html=True)

st.markdown('<div class="stitle">Catálogo territorial para completar el Excel</div>', unsafe_allow_html=True)
st.markdown('<div class="ibox">El catálogo territorial se consulta desde un archivo <b>.txt</b> público del repositorio GitHub. No necesitas subir un archivo local.</div>', unsafe_allow_html=True)
github_txt_url = st.text_input('URL del .txt territorial en GitHub', value=st.session_state.github_txt_url, key='github_txt_url_input')
st.session_state.github_txt_url = github_txt_url
if github_txt_url:
    try:
        df_cat = cargar_catalogo_github_txt(github_txt_url)
        lookup_cat, cols_cat = preparar_lookup_territorial(df_cat)
        st.session_state.catalogo_df = df_cat
        st.session_state.catalogo_lookup = lookup_cat
        st.session_state.catalogo_cols = cols_cat
        st.success(f'Catálogo cargado desde GitHub: {len(df_cat):,} filas')
        cols_detectadas = {k:v for k,v in cols_cat.items() if v}
        st.caption(f'Columnas detectadas: {cols_detectadas}')
    except Exception as e:
        st.error(f'No se pudo leer el catálogo territorial desde GitHub: {e}')

if btn:
    G=st.session_state.graph_G; eq_cfg=st.session_state.equipos_cfg; n_eq=len(eq_cfg); nombres=[e['nombre'] for e in eq_cfg]; n_clust=n_eq*2
    df_w=df.copy(); df_w['equipo']='sin_asignar'; df_w['jornada']='sin_asignar'; df_w['cluster_geo']=-1
    df_w['carga_pond']=df_w.apply(lambda r: r['viv']*p['factor_r'] if str(r.get('tipo_entidad','')).startswith('sec') else r['viv'], axis=1)
    df_w['encuestador']=0; df_w['dia_operativo']=0; df_w['dia_inicio']=0; df_w['dia_fin']=0; df_w['dist_base_m']=0.0
    prog=st.progress(0, 'Iniciando...')
    prog.progress(10,'KMeans y rebalanceo...')
    labels, ballog, cvini, cvfin = clustering_balanceado(df_w, n_clusters=n_clust, cv_objetivo=p['cv_objetivo']/100.0, max_iter=p['max_iter_bal'], k_vecinos=p['k_vecinos'])
    dfn = df_w.copy(); dfn['cluster_geo']=labels
    st.session_state.balance_log=ballog; st.session_state.cv_ini_bal=cvini; st.session_state.cv_fin_bal=cvfin
    centroides=np.array([dfn[dfn['cluster_geo']==c][['x','y']].mean().values for c in range(n_clust)])
    bx,by = dfn['x'].mean(), dfn['y'].mean()
    distc=np.sqrt((centroides[:,0]-bx)**2+(centroides[:,1]-by)**2); orden=np.argsort(distc)
    asig={}
    for i,(cj1,cj2) in enumerate(zip(orden[:n_eq], orden[n_eq:])):
        asig[cj1]=(nombres[i],'Jornada 1'); asig[cj2]=(nombres[i],'Jornada 2')
    dfn['equipo']=dfn['cluster_geo'].map(lambda c: asig.get(c, ('sin_asignar','sin_asignar'))[0])
    dfn['jornada']=dfn['cluster_geo'].map(lambda c: asig.get(c, ('sin_asignar','sin_asignar'))[1])
    for nombre_eq in nombres:
        for jornada in ['Jornada 1','Jornada 2']:
            mask=(dfn['equipo']==nombre_eq)&(dfn['jornada']==jornada)
            grp=dfn[mask].copy()
            if len(grp)==0: continue
            n_enc=next((e['enc'] for e in eq_cfg if e['nombre']==nombre_eq),3)
            ga=asignar_encuestadores_y_dias(grp, n_enc, p['dias_op'], p['viv_min'], p['viv_max'], 1)
            dfn.update(ga[['encuestador','dia_operativo','dia_inicio','dia_fin']])
    prog.progress(100,'Listo')
    resumen = dfn[dfn['equipo']!='sin_asignar'].groupby(['equipo','jornada']).agg(n_upms=('id_entidad','count'),viv_reales=('viv','sum'),carga_ponderada=('carga_pond','sum')).reset_index()
    st.session_state.df_plan=dfn; st.session_state.resumen_bal=resumen; st.session_state.resultados_generados=True
    st.session_state.editslog=[]; st.session_state.editcounter=0
    st.success('Planificación generada.')

if not st.session_state.resultados_generados:
    st.markdown('<div class="ibox">Presiona <b>Generar Planificación</b>.</div>', unsafe_allow_html=True)
    st.stop()

df_plan=st.session_state.df_plan
res_bal=st.session_state.resumen_bal
eq_cfg=st.session_state.equipos_cfg
nombres=[e['nombre'] for e in eq_cfg]
colormap={n:COLORES[i%len(COLORES)] for i,n in enumerate(nombres)}

j1n,j2n,mesnom=jornada_num_desde_mes(int(df['mes'].iloc[0]), st.session_state.mes_inicio_cal)

tabmapa,tabanalisis,tabedicion,tabreporte = st.tabs(['Mapa de Rutas','Análisis de Carga','Edición Manual','Reporte y Descarga'])

with tabmapa:
    st.markdown('<div class="stitle">Mapa del Operativo de Campo</div>', unsafe_allow_html=True)
    cc1,cc2=st.columns([1,3])
    with cc1:
        mj1=st.checkbox('Jornada 1',value=True)
        mj2=st.checkbox('Jornada 2',value=True)
        fnd=st.selectbox('Fondo',['CartoDB positron','OpenStreetMap','CartoDB darkmatter'])
    with cc2:
        m=folium.Map(location=[BASE_LAT,BASE_LON], zoom_start=8, tiles=fnd)
        folium.Marker([BASE_LAT,BASE_LON], popup='Base INEC GYE').add_to(m)
        for _,row in df_plan.iterrows():
            jor=row.get('jornada','')
            if jor=='Jornada 1' and not mj1: continue
            if jor=='Jornada 2' and not mj2: continue
            clr=colormap.get(row.get('equipo',''),'#888888')
            folium.CircleMarker(location=[row['lat'],row['lon']], radius=5, color=clr, fill=True, fill_color=clr, fill_opacity=.85, popup=f"ID {row.get('id_entidad','')}<br>Viv {int(row.get('viv',0))}<br>{row.get('equipo','')} · {jor}").add_to(m)
        st_folium(m, width=None, height=540)

with tabanalisis:
    st.markdown('<div class="stitle">Rebalanceo de Clusters</div>', unsafe_allow_html=True)
    cvini=st.session_state.get('cv_ini_bal'); cvfin=st.session_state.get('cv_fin_bal')
    if cvini is not None and cvfin is not None:
        mejora=cvini-cvfin
        st.markdown(f'<div class="balance-box"><b>CV inicial</b> <span style="color:#dc2626">{cvini:.1f}</span> &nbsp;&nbsp; <b>CV final</b> <span style="color:#059669">{cvfin:.1f}</span> &nbsp;&nbsp; <b>Mejora</b> {mejora:.1f} pp</div>', unsafe_allow_html=True)
    if res_bal is not None and len(res_bal)>0:
        fig=px.bar(res_bal, x='equipo', y='viv_reales', color='jornada', barmode='group', template='plotly_white', title='Viviendas por equipo y jornada', color_discrete_map={'Jornada 1':'#003B71','Jornada 2':'#059669'})
        st.plotly_chart(fig, use_container_width=True)
        st.dataframe(res_bal, use_container_width=True, hide_index=True)

with tabedicion:
    st.markdown('<div class="stitle">Edición Manual de la Planificación</div>', unsafe_allow_html=True)
    st.markdown('<div class="ibox">Aquí puedes ajustar manualmente equipo, jornada, encuestador y días operativos. Los cambios se reflejan en el mapa, análisis y Excel final.</div>', unsafe_allow_html=True)
    if st.session_state.editcounter > 0:
        st.markdown(f'<div class="edit-ok">{st.session_state.editcounter} ediciones manuales aplicadas en esta sesión.</div>', unsafe_allow_html=True)
    modoedit = st.radio('Modo de edición', ['Selección individual','Selección por lote'], horizontal=True)
    equiposdestino = nombres
    jornadasdestino = ['Jornada 1','Jornada 2']
    if modoedit == 'Selección individual':
        idsseleccionados = st.multiselect('UPMs / códigos a editar', sorted(df_plan['id_entidad'].astype(str).tolist()))
        nuevoequipo = st.selectbox('Nuevo equipo', ['Sin cambio'] + equiposdestino)
        nuevajornada = st.selectbox('Nueva jornada', ['Sin cambio'] + jornadasdestino)
        nuevoenc = st.number_input('Nuevo encuestador (0 sin cambio)', min_value=0, max_value=20, value=0)
        ed1, ed2 = st.columns(2)
        with ed1:
            nuevodiaini = st.number_input('Nuevo día inicio (0 sin cambio)', min_value=0, max_value=p['dias_op'], value=0)
        with ed2:
            nuevodiafin = st.number_input('Nuevo día fin (0 sin cambio)', min_value=0, max_value=p['dias_op'], value=0)
        if idsseleccionados:
            st.dataframe(df_plan[df_plan['id_entidad'].astype(str).isin(idsseleccionados)][['id_entidad','equipo','jornada','encuestador','dia_inicio','dia_fin','viv']], use_container_width=True, hide_index=True)
        if st.button('Aplicar cambios', type='primary', key='btnaplicarindividual', use_container_width=True):
            if idsseleccionados:
                masksel = df_plan['id_entidad'].astype(str).isin(idsseleccionados)
                idxedit = df_plan.index[masksel]
                logentry = {'tipo':'individual','upms':idsseleccionados.copy(),'cambios':{}}
                if nuevoequipo != 'Sin cambio': df_plan.loc[idxedit, 'equipo'] = nuevoequipo; logentry['cambios']['equipo']=nuevoequipo
                if nuevajornada != 'Sin cambio': df_plan.loc[idxedit, 'jornada'] = nuevajornada; logentry['cambios']['jornada']=nuevajornada
                if nuevoenc != 0: df_plan.loc[idxedit, 'encuestador'] = int(nuevoenc); logentry['cambios']['encuestador']=int(nuevoenc)
                if nuevodiaini != 0: df_plan.loc[idxedit, 'dia_inicio'] = int(nuevodiaini); df_plan.loc[idxedit, 'dia_operativo'] = int(nuevodiaini); logentry['cambios']['dia_inicio']=int(nuevodiaini)
                if nuevodiafin != 0: df_plan.loc[idxedit, 'dia_fin'] = int(nuevodiafin); logentry['cambios']['dia_fin']=int(nuevodiafin)
                st.session_state.df_plan = df_plan
                st.session_state.editslog.append(logentry)
                st.session_state.editcounter += 1
                resumen = df_plan.groupby(['equipo','jornada']).agg(n_upms=('id_entidad','count'),viv_reales=('viv','sum'),carga_ponderada=('carga_pond','sum')).reset_index()
                st.session_state.resumen_bal = resumen
                st.success(f'{len(idsseleccionados)} UPMs actualizadas.')
                st.rerun()
    else:
        loteequipoorig = st.selectbox('Equipo origen', ['Todos'] + sorted(df_plan['equipo'].astype(str).unique().tolist()))
        lotejornadaorig = st.selectbox('Jornada origen', ['Todas'] + sorted(df_plan['jornada'].astype(str).unique().tolist()))
        dfctx = df_plan.copy()
        if loteequipoorig != 'Todos': dfctx = dfctx[dfctx['equipo'] == loteequipoorig]
        if lotejornadaorig != 'Todas': dfctx = dfctx[dfctx['jornada'] == lotejornadaorig]
        st.dataframe(dfctx[['id_entidad','equipo','jornada','encuestador','dia_inicio','dia_fin','viv']].head(200), use_container_width=True, hide_index=True)
        loteequipodest = st.selectbox('Equipo destino', ['Sin cambio'] + equiposdestino)
        lotejornadadest = st.selectbox('Jornada destino', ['Sin cambio'] + jornadasdestino)
        loteencdest = st.number_input('Encuestador destino (0 sin cambio)', min_value=0, max_value=20, value=0)
        if st.button('Aplicar cambio masivo', type='primary', key='btnaplicarlote', use_container_width=True):
            if len(dfctx) > 0:
                idxlote = dfctx.index
                logentry = {'tipo':'lote','origen':f'{loteequipoorig}/{lotejornadaorig}','nupms':len(dfctx),'cambios':{}}
                if loteequipodest != 'Sin cambio': df_plan.loc[idxlote, 'equipo'] = loteequipodest; logentry['cambios']['equipo']=loteequipodest
                if lotejornadadest != 'Sin cambio': df_plan.loc[idxlote, 'jornada'] = lotejornadadest; logentry['cambios']['jornada']=lotejornadadest
                if loteencdest != 0: df_plan.loc[idxlote, 'encuestador'] = int(loteencdest); logentry['cambios']['encuestador']=int(loteencdest)
                st.session_state.df_plan = df_plan
                st.session_state.editslog.append(logentry)
                st.session_state.editcounter += 1
                resumen = df_plan.groupby(['equipo','jornada']).agg(n_upms=('id_entidad','count'),viv_reales=('viv','sum'),carga_ponderada=('carga_pond','sum')).reset_index()
                st.session_state.resumen_bal = resumen
                st.success(f'{len(dfctx)} UPMs actualizadas.')
                st.rerun()
    st.markdown('<div class="stitle">Historial de ediciones manuales</div>', unsafe_allow_html=True)
    if not st.session_state.editslog:
        st.markdown('<div class="ibox">Sin ediciones manuales en esta sesión.</div>', unsafe_allow_html=True)
    else:
        for i,loge in enumerate(reversed(st.session_state.editslog),1):
            if loge['tipo']=='individual':
                upmsstr=', '.join(loge['upms'][:5]) + (f" ... +{len(loge['upms'])-5} más" if len(loge['upms'])>5 else '')
                cambiosstr=', '.join(f'{k}: {v}' for k,v in loge['cambios'].items())
                st.markdown(f'<div class="edit-card"><b>{len(st.session_state.editslog)-i+1}</b> · Individual · UPMs: {upmsstr}<br><span style="font-size:11px">{cambiosstr}</span></div>', unsafe_allow_html=True)
            else:
                cambiosstr=', '.join(f'{k}: {v}' for k,v in loge['cambios'].items())
                st.markdown(f'<div class="edit-card"><b>{len(st.session_state.editslog)-i+1}</b> · Lote · {loge["nupms"]} UPMs desde {loge["origen"]}<br><span style="font-size:11px">{cambiosstr}</span></div>', unsafe_allow_html=True)

with tabreporte:
    st.markdown('<div class="stitle">Reporte y Descarga Excel</div>', unsafe_allow_html=True)
    st.markdown('<div class="ibox">El reporte se genera únicamente para el <b>mes operativo seleccionado</b>. Se incluyen solo la Jornada 1 y Jornada 2 de ese mes.</div>', unsafe_allow_html=True)
    if st.session_state.editcounter > 0:
        st.markdown(f'<div class="edit-ok">El Excel incluirá {st.session_state.editcounter} ediciones manuales.</div>', unsafe_allow_html=True)
    if res_bal is not None and len(res_bal)>0:
        tr=pd.DataFrame([{'equipo':'TOTAL','jornada':'—','n_upms':res_bal['n_upms'].sum(),'viv_reales':res_bal['viv_reales'].sum(),'carga_ponderada':res_bal['carga_ponderada'].sum()}])
        st.dataframe(pd.concat([res_bal,tr],ignore_index=True).rename(columns={'equipo':'Equipo','jornada':'Jornada','n_upms':'UPMs','viv_reales':'Viv.','carga_ponderada':'Carga pond.'}), use_container_width=True, hide_index=True)
    colsok=[c for c in ['id_entidad','tipo_entidad','viv','carga_pond','equipo','jornada','encuestador','dia_inicio','dia_fin'] if c in df_plan.columns]
    st.dataframe(df_plan[colsok].sort_values(['equipo','jornada','encuestador','dia_inicio']).reset_index(drop=True), use_container_width=True, height=280)
    jornadas_excel=[{'jornada_num':j1n,'jornada_nombre':'Jornada 1','fecha':None},{'jornada_num':j2n,'jornada_nombre':'Jornada 2','fecha':None}]
    if st.button('Generar Excel', use_container_width=True, type='primary'):
        try:
            excel_bytes=generar_excel(df_plan=st.session_state.df_plan, eq_cfg=eq_cfg, personal_info=st.session_state.personal_info, jornadas_activas=jornadas_excel, dias_op=p['dias_op'], catalogo_lookup=st.session_state.catalogo_lookup)
            fname=f'planificacion_J{j1n}_{j2n}_{mesnom}.xlsx'
            st.download_button(label=f'Descargar {fname}', data=excel_bytes, file_name=fname, mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', use_container_width=True)
            st.success('Excel listo.')
        except Exception as e:
            st.error(f'Error: {e}')
